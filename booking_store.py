from __future__ import annotations

import copy
import re
import uuid
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation
from email import policy
from email.parser import BytesParser
from pathlib import Path
from typing import TYPE_CHECKING, Any

from app.modules.booking.rules.registry import SUPPLIER_RULES, get_supplier_names
from app.shared.lazy_imports import lazy_module
from app_paths import RESOURCE_DIR, RUNTIME_DIR

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

openpyxl = lazy_module("openpyxl")
xlrd = lazy_module("xlrd")

BOOKING_TEMPLATE_NAME = "booking_template_zh.xlsx"


@dataclass
class BookingPreview:
    session_id: str
    supplier: str
    source_filename: str
    pack_filename: str
    rows: list[dict[str, Any]]
    columns: list[str]
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    detail_count: int = 0
    packadc_count: int = 0
    output_path: Path | None = None
    email_subject: str = ""
    mawb_no: str = ""
    purchaser: str = ""
    delivery_method: str = ""
    carrier_name: str = ""

    @property
    def can_generate(self) -> bool:
        return not self.errors and bool(self.rows)


@dataclass
class BookingEmailAttachment:
    filename: str
    path: Path
    content_type: str


def available_suppliers() -> list[str]:
    return get_supplier_names()


def safe_attachment_name(filename: str) -> str:
    clean = re.sub(r'[<>:"/\\|?*\r\n]+', "_", filename).strip()
    return clean or f"attachment_{uuid.uuid4().hex[:8]}"


def _is_excel_file(filename: str) -> bool:
    return filename.lower().endswith((".xls", ".xlsx"))


def _looks_like_source_name(filename: str) -> bool:
    lower = filename.lower()
    return _is_excel_file(filename) and (
        "ccixls" in lower
        or "packing" in lower
        or re.search(r"(^|[-_\s])inv[-_\s]", lower) is not None
        or re.search(r"(^|[-_\s])pl(?:[-_\s.]|$|\()", lower) is not None
        or re.search(r"pl(?:\(\d+\))?\.xlsx?$", lower) is not None
    )


def _looks_like_pack_name(filename: str) -> bool:
    lower = filename.lower()
    return _is_excel_file(filename) and ("packadcxls" in lower or re.search(r"(^|[-_\s])pak[-_\s]", lower) is not None)


def extract_booking_attachments_from_eml(eml_path: Path, target_dir: Path) -> tuple[Path | None, Path | None, list[str], str]:
    target_dir.mkdir(parents=True, exist_ok=True)
    message = BytesParser(policy=policy.default).parsebytes(eml_path.read_bytes())
    subject = str(message.get("subject") or "")
    warnings: list[str] = []
    ccixls_path: Path | None = None
    packadc_path: Path | None = None
    excel_attachments: list[tuple[str, Path]] = []

    for index, part in enumerate(message.walk(), start=1):
        filename = part.get_filename()
        if not filename:
            continue
        payload = part.get_payload(decode=True)
        if not payload:
            continue
        safe_name = safe_attachment_name(filename)
        target_path = target_dir / f"{index:03d}_{safe_name}"
        target_path.write_bytes(payload)
        if _is_excel_file(filename):
            excel_attachments.append((filename, target_path))
        if _looks_like_source_name(filename) and ccixls_path is None:
            ccixls_path = target_path
        elif _looks_like_pack_name(filename) and packadc_path is None:
            packadc_path = target_path

    if ccixls_path is None:
        if excel_attachments:
            picked_name, ccixls_path = excel_attachments[0]
            warnings.append(f"邮件附件中未找到 CCIXLS/INV/PL 命名的 Excel，已临时使用 {picked_name} 作为源文件。")
        else:
            warnings.append("邮件附件中未找到 CCIXLS/INV/PL Excel 附件。")
    return ccixls_path, packadc_path, warnings, subject


def extract_all_booking_attachments_from_eml(eml_path: Path, target_dir: Path) -> tuple[list[BookingEmailAttachment], str]:
    target_dir.mkdir(parents=True, exist_ok=True)
    message = BytesParser(policy=policy.default).parsebytes(eml_path.read_bytes())
    subject = str(message.get("subject") or "")
    attachments: list[BookingEmailAttachment] = []
    for index, part in enumerate(message.walk(), start=1):
        filename = part.get_filename()
        if not filename:
            continue
        payload = part.get_payload(decode=True)
        if not payload:
            continue
        safe_name = safe_attachment_name(filename)
        target_path = target_dir / f"{index:03d}_{safe_name}"
        target_path.write_bytes(payload)
        attachments.append(
            BookingEmailAttachment(
                filename=filename,
                path=target_path,
                content_type=part.get_content_type() or "application/octet-stream",
            )
        )
    return attachments, subject


def get_default_booking_template(rule: Any | None = None) -> Path:
    rule_candidates = [Path(item) for item in getattr(rule, "TEMPLATE_CANDIDATES", [])]
    candidates = [
        *rule_candidates,
        RESOURCE_DIR / BOOKING_TEMPLATE_NAME,
        Path.cwd() / BOOKING_TEMPLATE_NAME,
        Path(r"C:/Users/ac/Desktop/booking data from customer/booking_template_zh (1).xlsx"),
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    raise FileNotFoundError("未找到 booking_template_zh.xlsx 模板。")


def normalize_key(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.strip().lower()
    text = re.sub(r"[\s\r\n\t*．.。:：()（）/#\\-]+", "", text)
    return text


def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return f"{value:.0f}"
        return "{:.0f}".format(value)
    text = str(value).strip()
    if not text:
        return ""
    try:
        dec = Decimal(text)
    except InvalidOperation:
        return text
    if dec == dec.to_integral_value():
        return f"{dec:.0f}"
    return format(dec, "f").rstrip("0").rstrip(".")


def as_number(value: Any) -> int | float | None:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return int(value) if float(value).is_integer() else float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return 0
    try:
        number = Decimal(text)
    except InvalidOperation:
        return 0
    if number == number.to_integral_value():
        return int(number)
    return float(number)


def volume_cm_to_m3(value: Any) -> float | None:
    text = as_text(value).replace("×", "*").replace("x", "*").replace("X", "*")
    numbers = re.findall(r"\d+(?:\.\d+)?", text.replace(",", ""))
    if len(numbers) >= 3:
        volume = Decimal("1")
        for item in numbers[:3]:
            volume *= Decimal(item)
    elif len(numbers) == 1:
        volume = Decimal(numbers[0])
    else:
        return None
    return float(round(volume / Decimal("1000000"), 6))


def weight_to_kg(value: Any) -> int | float | None:
    values = value if isinstance(value, (list, tuple)) else [value, "KG"]
    amount = as_number(values[0] if values else 0) or 0
    unit = as_text(values[1] if len(values) > 1 else "KG").upper()
    if unit in {"KG", "KGS", "KILOGRAM", "KILOGRAMS", ""}:
        result = Decimal(str(amount))
    elif unit in {"G", "GS", "GRAM", "GRAMS"}:
        result = Decimal(str(amount)) / Decimal("1000")
    elif unit in {"MG", "MILLIGRAM", "MILLIGRAMS"}:
        result = Decimal(str(amount)) / Decimal("1000000")
    elif unit in {"LB", "LBS", "POUND", "POUNDS"}:
        result = Decimal(str(amount)) * Decimal("0.45359237")
    else:
        result = Decimal(str(amount))
    return int(result) if result == result.to_integral_value() else float(result)


def clean_value(value: Any, cleaner: str | None) -> Any:
    if cleaner == "suffix_0001":
        base = as_text(value).strip().rstrip(".．。").strip()
        return f"{base}-0001" if base else ""
    if cleaner == "join_dash":
        values = value if isinstance(value, (list, tuple)) else [value]
        parts = [as_text(item).strip() for item in values if as_text(item).strip()]
        return "-".join(parts)
    if cleaner == "join_dash_zfill4":
        values = value if isinstance(value, (list, tuple)) else [value]
        parts = [as_text(item).strip() for item in values if as_text(item).strip()]
        if len(parts) > 1 and parts[-1].isdigit():
            parts[-1] = parts[-1].zfill(4)
        return "-".join(parts)
    if cleaner == "as_text":
        return as_text(value)
    if cleaner == "as_number":
        return as_number(value)
    if cleaner == "volume_cm_to_m3":
        return volume_cm_to_m3(value)
    if cleaner == "weight_to_kg":
        return weight_to_kg(value)
    if cleaner == "letters_only":
        return "".join(re.findall(r"[A-Za-z]+", as_text(value))).upper()
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return f"{value:.0f}"
    return value


def extract_mawb_no(subject: str) -> str:
    match = re.search(r"MAWB\s*#\s*([A-Z0-9]+)", subject or "", flags=re.IGNORECASE)
    return match.group(1) if match else ""


def extract_mblno_from_text(text: str) -> str:
    match = re.search(
        r"(?i)\b(?:M\s*B\s*/?\s*L\s*NO|MBLNO|H\s*B\s*/?\s*L\s*NO|MAWB\s*#?|AIR\s+WAYBILL\s+NO)\s*[:：#*]*\s*([A-Z0-9][A-Z0-9-]*)",
        text or "",
    )
    return match.group(1).strip("-").upper() if match else ""


def extract_mblno_from_workbook(path: Path) -> str:
    try:
        if path.suffix.lower() == ".xls":
            book = xlrd.open_workbook(str(path), formatting_info=False)
            for sheet in book.sheets():
                for row_index in range(sheet.nrows):
                    row_text = " ".join(as_text(sheet.cell_value(row_index, col_index)) for col_index in range(sheet.ncols))
                    found = extract_mblno_from_text(row_text)
                    if found:
                        return found
        elif path.suffix.lower() == ".xlsx":
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    row_text = " ".join(as_text(value) for value in row)
                    found = extract_mblno_from_text(row_text)
                    if found:
                        return found
    except Exception:
        return ""
    return ""


def extract_booking_mblno(subject: str, source_path: Path, packadc_path: Path | None = None) -> str:
    return (
        extract_mblno_from_text(subject)
        or extract_mawb_no(subject)
        or extract_mblno_from_workbook(source_path)
        or (extract_mblno_from_workbook(packadc_path) if packadc_path else "")
    )


def normalize_tracking_no(value: str) -> str:
    return re.sub(r"[\s-]+", "", value or "").upper()


def classify_delivery_profile(mblno: str) -> tuple[str, str, str]:
    tracking_no = normalize_tracking_no(mblno)
    if not tracking_no:
        return "", "", ""
    if tracking_no.startswith("DIM"):
        return "LOCAL", "DIMERCO", ""
    if tracking_no.startswith("SF"):
        return "EXPRESS", "SF", ""
    if re.fullmatch(r"\d{10}", tracking_no):
        return "EXPRESS", "DHL", ""
    if re.fullmatch(r"\d{12}", tracking_no):
        return "EXPRESS", "FEDEX", ""
    if re.fullmatch(r"\d{15}|\d{20}|\d{22}", tracking_no):
        return "", "", f"MBLNO {mblno} 可能是 FedEx 长单号，请人工确认交付方式和运输公司。"
    return "", "", f"未能根据 MBLNO {mblno} 自动识别交付方式和运输公司。"


from app.modules.booking.mail_builder import (
    build_sil_fuca_warehouse_mail_body,
    build_sil_fuca_warehouse_mail_subject,
    extract_sil_warehouse_no,
    generate_sil_fuca_warehouse_eml,
    load_sil_fuca_warehouse_template,
    replace_mail_template_values,
    save_sil_fuca_warehouse_template_from_eml,
)
from app.modules.booking.excel_io import load_rows_from_workbook


def po_prefix(value: Any) -> str:
    text = as_text(value).upper()
    return text.split("-", 1)[0].strip()


def resolve_purchaser(detail_rows: list[dict[str, Any]], rule: Any) -> tuple[str, list[str]]:
    mapping = getattr(rule, "PURCHASER_BY_PO_PREFIX", {})
    if not mapping:
        return "", []
    po_column = getattr(rule, "PURCHASER_PO_COLUMN", "Customer PO")
    purchasers: dict[str, set[str]] = {}
    unknown: set[str] = set()
    for row in detail_rows:
        prefix = po_prefix(get_row_value(row, po_column))
        if not prefix:
            continue
        purchaser = mapping.get(prefix)
        if purchaser:
            purchasers.setdefault(purchaser, set()).add(prefix)
        else:
            unknown.add(prefix)
    warnings: list[str] = []
    if unknown:
        warnings.append(f"以下 PO 前缀未配置采购方：{', '.join(sorted(unknown))}")
    if len(purchasers) > 1:
        detail = "; ".join(f"{buyer}: {', '.join(sorted(prefixes))}" for buyer, prefixes in purchasers.items())
        warnings.append(f"同一份 CCIXLS 命中多个采购方，请检查：{detail}")
        return "", warnings
    if not purchasers:
        return "", warnings
    return next(iter(purchasers)), warnings


def get_row_value(row: dict[str, Any], column_name: str) -> Any:
    wanted = normalize_key(column_name)
    for key, value in row.items():
        if normalize_key(key) == wanted:
            return value
    return ""


def get_source_value(row: dict[str, Any], source_column: Any) -> Any:
    if isinstance(source_column, (list, tuple)):
        return [get_row_value(row, column_name) for column_name in source_column]
    return get_row_value(row, source_column)


def iter_source_columns(source_column: Any) -> list[str]:
    if isinstance(source_column, (list, tuple)):
        return [str(item) for item in source_column]
    return [str(source_column)]


def require_columns(headers: list[str], required: list[str], label: str) -> list[str]:
    present = {normalize_key(header) for header in headers}
    return [column for column in required if normalize_key(column) not in present]


def resolve_source_rows(source_path: Path, packadc_path: Path | None, rule: Any) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str], list[str]]:
    warnings: list[str] = []
    detail_result = load_rows_from_workbook(source_path, rule.SOURCE_SHEETS["detail"])
    if detail_result is None:
        raise ValueError("未找到 detail sheet。")
    detail_rows, detail_headers, detail_sheet_name = detail_result

    packadc_rows: list[dict[str, Any]] = []
    packadc_headers: list[str] = []
    packadc_candidates = rule.SOURCE_SHEETS.get("packadc", [])
    if packadc_candidates:
        internal_pack = load_rows_from_workbook(source_path, packadc_candidates)
        if internal_pack is not None and normalize_key(internal_pack[2]) != normalize_key(detail_sheet_name):
            packadc_rows, packadc_headers, _pack_sheet_name = internal_pack
        elif packadc_path and packadc_path.exists():
            external_pack = load_rows_from_workbook(packadc_path, packadc_candidates)
            if external_pack is None:
                external_pack = load_rows_from_workbook(packadc_path, ["detail"])
            if external_pack is None:
                warnings.append("已上传 PACKADCXLS 文件，但未找到可用明细 sheet。")
            else:
                packadc_rows, packadc_headers, _pack_sheet_name = external_pack
        else:
            warnings.append("源文件内没有 PACKADCXLS sheet，且未上传单独 PACKADCXLS 文件。")

    return detail_rows, packadc_rows, detail_headers, warnings


def build_booking_preview(
    *,
    session_id: str | None,
    supplier: str,
    source_path: Path,
    packadc_path: Path | None = None,
    email_subject: str = "",
) -> BookingPreview:
    if supplier not in SUPPLIER_RULES:
        raise ValueError(f"不支持的供应商：{supplier}")
    rule = SUPPLIER_RULES[supplier]
    session_id = session_id or uuid.uuid4().hex[:12]
    warnings: list[str] = []
    if not _looks_like_source_name(source_path.name):
        warnings.append("源文件名不像 CCIXLS/INV 文件，请确认是否上传了正确的客户源文件。")

    detail_rows, packadc_rows, detail_headers, source_warnings = resolve_source_rows(source_path, packadc_path, rule)
    warnings.extend(source_warnings)

    optional_source_columns = {normalize_key(column) for column in getattr(rule, "OPTIONAL_SOURCE_COLUMNS", set())}
    detail_required: list[str] = []
    for source_alias, source_column, _cleaner in rule.COLUMN_MAP.values():
        if source_alias == "detail":
            detail_required.extend(
                column
                for column in iter_source_columns(source_column)
                if normalize_key(column) not in optional_source_columns
            )
    missing = require_columns(detail_headers, detail_required, "detail")
    if missing:
        mawb_no = extract_booking_mblno(email_subject, source_path, packadc_path)
        delivery_method, carrier_name, delivery_warning = classify_delivery_profile(mawb_no)
        if delivery_warning:
            warnings.append(delivery_warning)
        return BookingPreview(
            session_id=session_id,
            supplier=supplier,
            source_filename=source_path.name,
            pack_filename=packadc_path.name if packadc_path else "",
            rows=[],
            columns=[],
            warnings=warnings,
            errors=[f"detail 缺少必要字段：{', '.join(missing)}"],
            detail_count=len(detail_rows),
            packadc_count=len(packadc_rows),
            email_subject=email_subject,
            mawb_no=mawb_no,
            delivery_method=delivery_method,
            carrier_name=carrier_name,
        )

    prepare_rows = getattr(rule, "prepare_rows", None)
    if callable(prepare_rows):
        detail_rows, prepare_warnings = prepare_rows(detail_rows)
        warnings.extend(prepare_warnings)

    pack_extras, pack_warnings = rule.post_process(detail_rows, packadc_rows)
    warnings.extend(pack_warnings)
    purchaser, purchaser_warnings = resolve_purchaser(detail_rows, rule)
    warnings.extend(purchaser_warnings)
    mawb_no = extract_booking_mblno(email_subject, source_path, packadc_path)
    if not mawb_no:
        warnings.append("邮件主题和源文件中未识别到 MBLNO/MAWB#，请检查源文件格式。")
    delivery_method, carrier_name, delivery_warning = classify_delivery_profile(mawb_no)
    if delivery_warning:
        warnings.append(delivery_warning)

    columns = list(rule.COLUMN_MAP.keys())
    for column in rule.CONSTANTS:
        if column not in columns:
            columns.append(column)
    for column in getattr(rule, "QUANTITY_COPY_COLUMNS", []):
        if column not in columns:
            columns.append(column)
    zero_columns = list(getattr(rule, "ZERO_AFTER_FIRST_COLUMNS", []))
    for column in zero_columns:
        if column not in columns:
            columns.append(column)

    volume_column = zero_columns[0] if len(zero_columns) > 0 else ""
    pallet_column = zero_columns[1] if len(zero_columns) > 1 else ""
    carton_column = zero_columns[2] if len(zero_columns) > 2 else ""
    qty_column = next((key for key in rule.COLUMN_MAP if normalize_key(key) == normalize_key("数量")), "数量")
    part_no_column = next((key for key in rule.COLUMN_MAP if normalize_key(key) == normalize_key("启益料号")), "启益料号")

    output_rows: list[dict[str, Any]] = []
    for index, detail_row in enumerate(detail_rows):
        extra = pack_extras[index] if index < len(pack_extras) else {}
        output: dict[str, Any] = {}
        for target_column, (source_alias, source_column, cleaner) in rule.COLUMN_MAP.items():
            source = extra if source_alias in {"packadc_match", "row_extra"} else detail_row
            output[target_column] = clean_value(get_source_value(source, source_column), cleaner)
        for column, value in rule.CONSTANTS.items():
            output[column] = value

        quantity = output.get("数量", output.get(qty_column, 0) if qty_column else 0)
        for column in getattr(rule, "QUANTITY_COPY_COLUMNS", []):
            output[column] = quantity
        output.setdefault("最小包装数", quantity)
        output.setdefault("每箱标准数", quantity)
        for column in getattr(rule, "FALLBACK_TO_QUANTITY_COLUMNS", []):
            if output.get(column) in ("", None, 0):
                output[column] = quantity
        for column in zero_columns:
            output[column] = 0 if index > 0 else ""
        if not output.get("批次", ""):
            output["批次"] = "0"
        if not quantity:
            warnings.append(f"第 {index + 1} 行数量为空或 0。")
        part_no = str(output.get("启益料号", output.get(part_no_column, "")) if part_no_column else output.get("启益料号", ""))
        if "e+" in part_no.lower() or len(part_no) < 6:
            warnings.append(f"第 {index + 1} 行启益料号疑似异常：{part_no}")
        output_rows.append(output)

    if supplier == "SIL-FUCA" and output_rows:
        first_extra = pack_extras[0] if pack_extras else {}
        box_count = int(as_number(first_extra.get("Box Count", 0)) or 0)
        if carton_column:
            output_rows[0][carton_column] = box_count
        if volume_column:
            output_rows[0][volume_column] = round(box_count * 0.01, 2)
        if pallet_column:
            output_rows[0][pallet_column] = ""

    return BookingPreview(
        session_id=session_id,
        supplier=supplier,
        source_filename=source_path.name,
        pack_filename=packadc_path.name if packadc_path else "",
        rows=output_rows,
        columns=columns,
        warnings=warnings,
        errors=[],
        detail_count=len(detail_rows),
        packadc_count=len(packadc_rows),
        email_subject=email_subject,
        mawb_no=mawb_no,
        purchaser=purchaser,
        delivery_method=delivery_method,
        carrier_name=carrier_name,
    )


def _header_column_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        if value is None:
            continue
        key = normalize_key(value)
        if key:
            mapping[key] = col
    return mapping


def _find_total_row(ws: Worksheet, start_row: int) -> int:
    for row in range(start_row, ws.max_row + 1):
        value = ws.cell(row, 1).value
        if isinstance(value, str) and normalize_key(value) == "total":
            return row
    return start_row


def _set_merged_value(ws: Worksheet, cell_ref: str, value: Any) -> None:
    ws[cell_ref].value = value


def _copy_row_style(ws: Worksheet, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy.copy(source.alignment)
        if source.protection:
            target.protection = copy.copy(source.protection)


def _unmerge_ranges_intersecting_rows(ws: Worksheet, start_row: int, end_row: int, style_source_row: int) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.max_row < start_row or merged_range.min_row > end_row:
            continue
        affected_rows = range(max(merged_range.min_row, start_row), min(merged_range.max_row, end_row) + 1)
        affected_cols = range(merged_range.min_col, merged_range.max_col + 1)
        ws.unmerge_cells(str(merged_range))
        for row in affected_rows:
            ws.row_dimensions[row].height = ws.row_dimensions[style_source_row].height
            for col in affected_cols:
                source = ws.cell(style_source_row, col)
                target = ws.cell(row, col)
                if source.has_style:
                    target._style = copy.copy(source._style)
                if source.number_format:
                    target.number_format = source.number_format
                if source.alignment:
                    target.alignment = copy.copy(source.alignment)
                if source.protection:
                    target.protection = copy.copy(source.protection)


def write_booking_workbook(preview: BookingPreview, output_dir: Path | None = None, template_path: Path | None = None) -> Path:
    if not preview.can_generate:
        raise ValueError("当前预览有错误，不能生成 booking form。")
    rule = SUPPLIER_RULES[preview.supplier]
    template_path = template_path or get_default_booking_template(rule)
    output_dir = output_dir or (RUNTIME_DIR / "outputs")
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]
    _set_merged_value(ws, "K4", preview.purchaser)
    _set_merged_value(ws, "U4", preview.mawb_no)
    if preview.delivery_method:
        _set_merged_value(ws, "K5", preview.delivery_method)
    if preview.carrier_name:
        _set_merged_value(ws, "Q5", preview.carrier_name)
    _set_merged_value(ws, "G6", "NA")
    _set_merged_value(ws, "K6", "NA")
    _set_merged_value(ws, "U6", "NA")
    header_map = _header_column_map(ws, rule.HEADER_ROW)
    required_targets = set(preview.columns)
    missing_targets = [column for column in required_targets if normalize_key(column) not in header_map]
    if missing_targets:
        raise ValueError(f"booking 模板缺少目标列：{', '.join(missing_targets)}")

    data_start = rule.DATA_START_ROW
    data_end = rule.DATA_END_ROW
    total_row = _find_total_row(ws, data_start)
    available = max(0, min(data_end, total_row - 1) - data_start + 1)
    needed = len(preview.rows)
    if needed > available:
        extra_rows = needed - available
        insert_at = total_row
        ws.insert_rows(insert_at, extra_rows)
        for offset in range(extra_rows):
            _copy_row_style(ws, data_end, insert_at + offset)
            ws.cell(insert_at + offset, 1).value = available + offset + 1

    final_data_end = data_start + max(needed, available) - 1
    _unmerge_ranges_intersecting_rows(ws, data_start, final_data_end, data_end)
    for row in range(data_start, final_data_end + 1):
        for col in range(2, ws.max_column + 1):
            ws.cell(row, col).value = None

    text_targets = {normalize_key(column) for column in getattr(rule, "TEXT_TARGET_COLUMNS", set())}
    for index, row_data in enumerate(preview.rows):
        excel_row = data_start + index
        ws.cell(excel_row, 1).value = index + 1
        for target_column, value in row_data.items():
            col = header_map.get(normalize_key(target_column))
            if not col:
                continue
            cell = ws.cell(excel_row, col)
            cell.value = value
            if normalize_key(target_column) in text_targets:
                cell.number_format = "@"

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    output_path = output_dir / f"booking_template_zh_{timestamp}.xlsx"
    wb.save(output_path)
    preview.output_path = output_path
    return output_path
