from __future__ import annotations

import argparse
import re
import zipfile
from copy import copy
from dataclasses import dataclass
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import TYPE_CHECKING, Any, Dict, Iterable, List, Optional, Sequence, Tuple
from xml.etree import ElementTree as ET

from app.shared.lazy_imports import lazy_module
from app.shared.performance import cached_file_result
from app_paths import RESOURCE_DIR

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

openpyxl = lazy_module("openpyxl")
pd = lazy_module("pandas")

XML_NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
MASTER_HEADER_ROW = 3
MASTER_DATA_START_ROW = 4
PALLET_UNIT = Decimal("63.78")
CARTON_UNIT = Decimal("10")
DELIVERY_UNIT = Decimal("239.16")
EXPRESS_UNIT = Decimal("30")
DEFAULT_PRICE_BOOK_PATH = Path(
    r"C:\Users\ac\Desktop\伟创力账单\福永伟创力账单\FY中港运输标路线final price.xlsx"
)
FALLBACK_PRICE_BOOK_PATH = RESOURCE_DIR / "sample_price.xlsx"
FIXED_DEFAULTS = {
    "伟创力事业部": "伟创力福永",
    "伟创力公司": "Fuyong",
    "伟创力厂房号": "813",
    "收货/发货": "Out 发货",
    "整车/拼车": "整车",
    "始发地\n(城市)": "深圳福永",
    "目的地\n(城市)": "香港",
    "币种": "HKD",
}
MONEY_COLUMNS = {
    "运费",
    "香港无缝费",
    "进仓费",
    "停车费",
    "多点送货费",
    "机场费",
    "租柜费",
    "装卸费",
    "传真快递费",
    "应收合计-不含税",
    "应收合计-含税",
}


def excel_serial_to_datetime(value: float) -> datetime:
    return datetime(1899, 12, 30) + timedelta(days=float(value))


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\u3000", " ")
    return "".join(text.split()).lower()


def parse_amount(raw: object) -> Decimal:
    if raw is None:
        return Decimal("0.00")
    cleaned = re.sub(r"[^0-9.\-]", "", str(raw).strip())
    if not cleaned:
        return Decimal("0.00")
    return Decimal(cleaned).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def make_excel_compatible_copy(path: Path) -> Path:
    workbook_path = path
    if path.suffix.lower() == ".xls":
        workbook_path = path.with_suffix(".xlsx")
        workbook_path.write_bytes(path.read_bytes())
    return workbook_path


def decimal_to_float(value: Decimal) -> float:
    return float(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def quantized(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def slt_sort_key(value: str) -> tuple[int, str]:
    text = str(value or "").strip().upper()
    match = re.search(r"(\d+)$", text)
    return (int(match.group(1)) if match else 10**9, text)


def col_letters_to_index(col_ref: str) -> int:
    value = 0
    for ch in col_ref:
        if ch.isalpha():
            value = value * 26 + ord(ch.upper()) - 64
    return value


def extract_cell_value(cell: ET.Element, shared_strings: Sequence[str]) -> Optional[str]:
    cell_type = cell.attrib.get("t")
    if cell_type == "s":
        value_node = cell.find("m:v", XML_NS)
        return shared_strings[int(value_node.text)] if value_node is not None else None
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell.iterfind(".//m:t", XML_NS))
    value_node = cell.find("m:v", XML_NS)
    return value_node.text if value_node is not None else None


def read_xlsx_rows(path: Path) -> Dict[int, Dict[int, object]]:
    with zipfile.ZipFile(path) as archive:
        shared_strings: List[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            shared_root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in shared_root.findall("m:si", XML_NS):
                shared_strings.append(
                    "".join(node.text or "" for node in item.iterfind(".//m:t", XML_NS))
                )

        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels_root}
        first_sheet = workbook_root.find("m:sheets/m:sheet", XML_NS)
        if first_sheet is None:
            raise ValueError(f"{path.name} 没有可读取的工作表")

        rel_id = first_sheet.attrib[
            "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
        ]
        sheet_path = "xl/" + rel_map[rel_id].lstrip("/")
        sheet_root = ET.fromstring(archive.read(sheet_path))

    rows: Dict[int, Dict[int, object]] = {}
    for cell in sheet_root.findall(".//m:sheetData/m:row/m:c", XML_NS):
        match = re.match(r"([A-Z]+)(\d+)", cell.attrib["r"])
        if not match:
            continue
        col = col_letters_to_index(match.group(1))
        row = int(match.group(2))
        rows.setdefault(row, {})[col] = extract_cell_value(cell, shared_strings)
    return rows


def find_value_beside_label(rows: Dict[int, Dict[int, object]], keywords: Iterable[str]) -> Optional[str]:
    normalized_keywords = [normalize_text(item) for item in keywords]
    for row_idx in sorted(rows):
        row = rows[row_idx]
        for col_idx in sorted(row):
            cell_text = normalize_text(row[col_idx])
            if any(keyword in cell_text for keyword in normalized_keywords):
                for next_col in range(col_idx + 1, max(row) + 1):
                    candidate = row.get(next_col)
                    if candidate not in (None, ""):
                        return str(candidate).strip()
    return None


def normalize_invoice_fee_name(text: object) -> str:
    raw = str(text or "")
    replacements = {
        "登記費": "登记费",
        "登記": "登记",
        "停車費": "停车费",
        "派送費": "派送费",
        "裝卸費": "装卸费",
        "機場附加費": "机场附加费",
        "运費": "运费",
        "運費": "运费",
        "無縫": "无缝",
        "快遞費/郵費": "快递费/邮费",
        "快遞費": "快递费",
        "郵費": "邮费",
        "費": "费",
    }
    for src, dest in replacements.items():
        raw = raw.replace(src, dest)
    return normalize_text(raw)


def maybe_correct_fee_name(fee_name: str, description: str, amount: Decimal) -> Tuple[str, Optional[str]]:
    normalized = normalize_invoice_fee_name(fee_name)
    description = str(description or "").strip()

    # Manually entered invoices sometimes type the fixed delivery charge 239.16
    # as 装卸费. When that happens, treat it as 派送费 instead of failing the split.
    if "装卸费" in normalized and amount == Decimal("239.16") and description:
        corrected = normalize_invoice_fee_name("派送费")
        reason = "金额为 239.16 且带有送货描述，疑似把派送费录成了装卸费"
        return corrected, reason

    return normalized, None


def normalize_truck_type(raw: Optional[str]) -> Optional[str]:
    if not raw:
        return None
    compact = normalize_text(raw)
    if "45" in compact:
        return "45尺"
    ton_match = re.search(r"(\d+)\s*t", compact)
    if ton_match:
        return f"{ton_match.group(1)}T"
    digit_match = re.search(r"(\d+)", compact)
    if digit_match:
        if "尺" in raw:
            return f"{digit_match.group(1)}尺"
        if "吨" in raw or "ton" in compact:
            return f"{digit_match.group(1)}T"
    return raw.strip()


def normalize_route_city(raw: Optional[str]) -> str:
    value = str(raw or "").strip().lower()
    if any(token in value for token in ["福永", "深圳"]):
        return "福永"
    if "香港" in value:
        return "香港"
    return str(raw or "").strip()


def normalize_price_truck_type(raw: Optional[str]) -> str:
    value = str(raw or "").strip().upper()
    if any(token in value for token in ["45HQ", "40HQ", "45尺"]):
        return "45尺"
    return normalize_truck_type(value) or value


def load_price_reference() -> Dict[tuple[str, str, str], Decimal]:
    path = DEFAULT_PRICE_BOOK_PATH if DEFAULT_PRICE_BOOK_PATH.exists() else FALLBACK_PRICE_BOOK_PATH
    if not path.exists():
        return {}

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    mapping: Dict[tuple[str, str, str], Decimal] = {}
    current_origin = ""
    current_destination = ""
    for row in range(3, ws.max_row + 1):
        origin = ws.cell(row, 3).value
        destination = ws.cell(row, 4).value
        truck_type = ws.cell(row, 5).value
        freight = ws.cell(row, 10).value
        if origin:
            current_origin = normalize_route_city(origin)
        if destination:
            current_destination = normalize_route_city(destination)
        if truck_type and freight not in (None, ""):
            normalized_truck = normalize_price_truck_type(truck_type)
            mapping[(current_origin, current_destination, normalized_truck)] = quantized(Decimal(str(freight)))
    return mapping


def normalize_delivery_fee_note(description: str) -> str:
    text = str(description or "").strip().lower()
    if not text:
        return ""
    text = re.sub(r"[0-9xX]+\s*板", "", text)
    text = re.sub(r"[0-9xX]+\s*箱", "", text)
    text = re.sub(r"[0-9]+\s*件", "", text)
    text = re.sub(r"[0-9.]+\s*kg", "", text)
    text = text.replace("派送", "").replace("送货", "").replace("送", "").replace("交", "")
    text = re.sub(r"[\s,，。.:：;；/()（）\-]+", "", text)
    return text


def parse_loading_count_hint(description: str) -> tuple[Decimal, Decimal]:
    text = str(description or "")
    pallet_count = Decimal("0.00")
    carton_count = Decimal("0.00")
    for match in re.finditer(r"(\d+(?:\.\d+)?)\s*([板箱])", text):
        count = Decimal(match.group(1))
        unit = match.group(2)
        if unit == "板":
            pallet_count += count
        elif unit == "箱":
            carton_count += count
    return pallet_count, carton_count


def build_fee_line_validation_issues(fee_items: Sequence["InvoiceFeeItem"]) -> List[Dict[str, Any]]:
    issues: List[Dict[str, Any]] = []
    loading_amount_mismatches: List[str] = []

    for item in fee_items:
        if "装卸费" not in item.normalized_fee_name:
            continue

        pallet_hint, carton_hint = parse_loading_count_hint(item.description)
        has_loading_hint = pallet_hint > 0 or carton_hint > 0
        if has_loading_hint:
            expected_amount = quantized(pallet_hint * PALLET_UNIT + carton_hint * CARTON_UNIT)
            if expected_amount != item.amount:
                loading_amount_mismatches.append(
                    f"第 {item.row_number} 行说明「{item.description or '-'}」应为 "
                    f"{decimal_to_float(expected_amount):.2f}，实际为 {decimal_to_float(item.amount):.2f}"
                )

    if loading_amount_mismatches:
        issues.append(
            {
                "type": "loading_fee_note_amount_mismatch",
                "title": "装卸费说明与金额不一致",
                "detail": "；".join(loading_amount_mismatches[:10])
                + "。请检查是否把板数按箱数收费，或把其他费用录进了装卸费。",
            }
        )

    return issues


def find_duplicate_delivery_fee_notes(fee_items: Sequence["InvoiceFeeItem"]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[InvoiceFeeItem]] = {}
    for item in fee_items:
        if "派送费" not in item.normalized_fee_name:
            continue
        normalized_note = normalize_delivery_fee_note(item.description)
        if not normalized_note:
            continue
        grouped.setdefault(normalized_note, []).append(item)

    duplicates: List[Dict[str, Any]] = []
    for _, items in grouped.items():
        if len(items) < 2:
            continue
        duplicates.append(
            {
                "normalized_note": normalize_delivery_fee_note(items[0].description),
                "display_note": items[0].description,
                "count": len(items),
                "rows": [item.row_number for item in items],
                "amounts": [decimal_to_float(item.amount) for item in items],
            }
        )
    duplicates.sort(key=lambda item: (-item["count"], item["display_note"]))
    return duplicates


def is_container_truck(raw_truck_type: Optional[str], normalized_truck_type: Optional[str]) -> bool:
    raw = str(raw_truck_type or "")
    normalized = str(normalized_truck_type or "")
    return any(token in raw for token in ["柜", "45尺"]) or "45尺" in normalized


def is_hong_kong_terminal(destination_city: Optional[str]) -> bool:
    destination = str(destination_city or "").strip().lower()
    return any(token in destination for token in ["码头", "terminal", "貨櫃碼頭", "货柜码头"])


def build_validation_issues(
    *,
    freight: Decimal,
    seamless_fee: Decimal,
    container_rental_fee: Decimal,
    multi_point_delivery_fee: Decimal,
    airport_fee: Decimal,
    express_fee: Decimal,
    raw_truck_type: Optional[str],
    normalized_truck_type: Optional[str],
    origin_city: Optional[str],
    destination_city: Optional[str],
    price_reference: Dict[tuple[str, str, str], Decimal],
) -> List[Dict[str, Any]]:
    issues: List[Dict[str, Any]] = []

    if seamless_fee <= Decimal("0.00"):
        issues.append(
            {
                "type": "missing_seamless_fee",
                "title": "缺少无缝",
                "detail": "这张账单没有无缝费用，请确认是否漏录。",
            }
        )

    if (
        is_container_truck(raw_truck_type, normalized_truck_type)
        and not is_hong_kong_terminal(destination_city)
        and container_rental_fee <= Decimal("0.00")
    ):
        issues.append(
            {
                "type": "missing_container_rental_fee",
                "title": "疑似缺少租柜费",
                "detail": "柜车且目的地不是香港码头，但账单中没有租柜费，请重点检查。",
            }
        )

    for fee_value, fee_type, title in [
        (multi_point_delivery_fee, "delivery_multiple_check", "派送费不是 239.16 的整数倍"),
        (airport_fee, "airport_multiple_check", "机场附加费不是 239.16 的整数倍"),
    ]:
        if fee_value > Decimal("0.00") and (fee_value % DELIVERY_UNIT) != Decimal("0.00"):
            issues.append(
                {
                    "type": fee_type,
                    "title": title,
                    "detail": f"当前金额为 {fee_value}，按规则应为 239.16 的整数倍，请检查是否录错或漏录。",
                }
            )

    if express_fee > Decimal("0.00") and (express_fee % EXPRESS_UNIT) != Decimal("0.00"):
        issues.append(
            {
                "type": "express_multiple_check",
                "title": "快递费不是 30 的整数倍",
                "detail": f"当前金额为 {express_fee}，按报价应为 30 的整数倍，请检查是否录错或漏录。",
            }
        )

    route_key = (
        normalize_route_city(origin_city),
        normalize_route_city(destination_city),
        normalize_price_truck_type(normalized_truck_type or raw_truck_type),
    )
    quoted_freight = price_reference.get(route_key)
    if quoted_freight is not None and abs(freight - quoted_freight) > Decimal("0.05"):
        issues.append(
            {
                "type": "freight_price_mismatch",
                "title": "运费与报价不一致",
                "detail": f"当前运费为 {freight}，报价参考为 {quoted_freight}，请检查录入是否正确。",
            }
        )

    return issues


@dataclass
class InvoiceFeeItem:
    row_number: int
    fee_name: str
    description: str
    amount: Decimal
    currency: str
    normalized_fee_name: str
    corrected_fee_name: str = ""
    suspected_name_error: bool = False

    def to_preview(self) -> Dict[str, Any]:
        return {
            "row_number": self.row_number,
            "fee_name": self.fee_name,
            "description": self.description,
            "amount": decimal_to_float(self.amount),
            "currency": self.currency,
            "normalized_fee_name": self.normalized_fee_name,
            "corrected_fee_name": self.corrected_fee_name,
            "suspected_name_error": self.suspected_name_error,
        }


@dataclass
class LoadingBreakdown:
    total_amount: Decimal
    pallet_count: int
    carton_count: int
    pallet_only_amount: Decimal
    remark: str
    exact_split: bool


@dataclass
class DeliveryNoteLine:
    delivery_no: str
    invoice_number: str
    delivery_time: str
    warehouse_so: str
    company_name: str
    area: str
    address: str
    total_pieces: Decimal
    pallet_count: Decimal
    billing_pallet_count: Decimal
    carton_count: Decimal
    gross_weight: Decimal

    def to_preview(self) -> Dict[str, Any]:
        return {
            "delivery_no": self.delivery_no,
            "invoice_number": self.invoice_number,
            "delivery_time": self.delivery_time,
            "warehouse_so": self.warehouse_so,
            "company_name": self.company_name,
            "area": self.area,
            "address": self.address,
            "total_pieces": decimal_to_float(self.total_pieces),
            "pallet_count": decimal_to_float(self.pallet_count),
            "billing_pallet_count": decimal_to_float(self.billing_pallet_count),
            "carton_count": decimal_to_float(self.carton_count),
            "gross_weight": decimal_to_float(self.gross_weight),
        }


@dataclass
class SourceTanRecord:
    customer_order_no: str
    tan_no: str
    description: str
    carton_count: Decimal
    pallet_count: Decimal
    loose_carton_count: Decimal
    gross_weight: Decimal
    row_count: int

    def to_preview(self) -> Dict[str, Any]:
        return {
            "customer_order_no": self.customer_order_no,
            "tan_no": self.tan_no,
            "description": self.description,
            "carton_count": decimal_to_float(self.carton_count),
            "pallet_count": decimal_to_float(self.pallet_count),
            "loose_carton_count": decimal_to_float(self.loose_carton_count),
            "gross_weight": decimal_to_float(self.gross_weight),
            "row_count": self.row_count,
        }


def split_loading(total_amount: Decimal) -> LoadingBreakdown:
    total_cents = int((quantized(total_amount) * 100).to_integral_value())
    pallet_cents = int((PALLET_UNIT * 100).to_integral_value())
    carton_cents = int((CARTON_UNIT * 100).to_integral_value())

    if total_cents % pallet_cents == 0:
        pallet_count = total_cents // pallet_cents
        return LoadingBreakdown(
            total_amount=quantized(total_amount),
            pallet_count=pallet_count,
            carton_count=0,
            pallet_only_amount=quantized(PALLET_UNIT * pallet_count),
            remark=str(pallet_count),
            exact_split=True,
        )

    pallet_count = None
    carton_count = None
    for candidate_pallets in range(total_cents // pallet_cents, -1, -1):
        remainder = total_cents - candidate_pallets * pallet_cents
        if remainder >= 0 and remainder % carton_cents == 0:
            pallet_count = candidate_pallets
            carton_count = remainder // carton_cents
            break
    if pallet_count is None or carton_count is None:
        raise ValueError(
            f"装卸费 {total_amount} 无法按 63.78/板 和 10/箱 拆分，请检查单页账单。"
        )
    return LoadingBreakdown(
        total_amount=quantized(total_amount),
        pallet_count=pallet_count,
        carton_count=carton_count,
        pallet_only_amount=quantized(PALLET_UNIT * pallet_count),
        remark=f"{pallet_count}板{carton_count}箱",
        exact_split=False,
    )


@dataclass
class InvoiceData:
    invoice_path: Path
    invoice_number: str
    customer_order_no: str
    job_date: datetime
    truck_type: Optional[str]
    origin_city: Optional[str]
    destination_city: Optional[str]
    currency: str
    freight: Decimal
    seamless_fee: Decimal
    registration_fee: Decimal
    parking_fee: Decimal
    inbound_fee: Decimal
    multi_point_delivery_fee: Decimal
    airport_fee: Decimal
    container_rental_fee: Decimal
    express_fee: Decimal
    loading: LoadingBreakdown
    fee_items: List[InvoiceFeeItem]
    suspected_fee_name_errors: List[Dict[str, Any]]
    validation_issues: List[Dict[str, Any]]
    quoted_freight: Optional[Decimal]
    delivery_note_lines: List[DeliveryNoteLine]
    source_tan_records: List[SourceTanRecord]
    delivery_validation_issues: List[Dict[str, Any]]

    def summary_preview(self) -> Dict[str, Any]:
        delivery_total_pieces = sum((item.total_pieces for item in self.delivery_note_lines), Decimal("0.00"))
        delivery_total_weight = sum((item.gross_weight for item in self.delivery_note_lines), Decimal("0.00"))
        delivery_total_pallets = sum((item.pallet_count for item in self.delivery_note_lines), Decimal("0.00"))
        source_total_cartons = sum((item.carton_count for item in self.source_tan_records), Decimal("0.00"))
        source_total_weight = sum((item.gross_weight for item in self.source_tan_records), Decimal("0.00"))
        source_total_pallets = sum((item.pallet_count for item in self.source_tan_records), Decimal("0.00"))
        source_total_loose_cartons = sum((item.loose_carton_count for item in self.source_tan_records), Decimal("0.00"))
        duplicate_delivery_fee_notes = find_duplicate_delivery_fee_notes(self.fee_items)
        return {
            "invoice_name": self.invoice_path.name,
            "invoice_number": self.invoice_number,
            "customer_order_no": self.customer_order_no,
            "job_date": self.job_date.strftime("%Y-%m-%d"),
            "truck_type": self.truck_type or "",
            "origin_city": self.origin_city or "",
            "destination_city": self.destination_city or "",
            "currency": self.currency,
            "freight": decimal_to_float(self.freight),
            "seamless_fee": decimal_to_float(self.seamless_fee),
            "registration_fee": decimal_to_float(self.registration_fee),
            "parking_fee": decimal_to_float(self.parking_fee),
            "inbound_fee": decimal_to_float(self.inbound_fee),
            "multi_point_delivery_fee": decimal_to_float(self.multi_point_delivery_fee),
            "airport_fee": decimal_to_float(self.airport_fee),
            "container_rental_fee": decimal_to_float(self.container_rental_fee),
            "express_fee": decimal_to_float(self.express_fee),
            "loading_fee": decimal_to_float(self.loading.total_amount),
            "loading_remark": self.loading.remark,
            "suspected_fee_name_error_count": len(self.suspected_fee_name_errors),
            "suspected_fee_name_errors": self.suspected_fee_name_errors,
            "validation_issue_count": len(self.validation_issues),
            "validation_issues": self.validation_issues,
            "quoted_freight": decimal_to_float(self.quoted_freight) if self.quoted_freight is not None else "",
            "delivery_note_lines": [item.to_preview() for item in self.delivery_note_lines],
            "source_tan_records": [item.to_preview() for item in self.source_tan_records],
            "delivery_total_pieces": decimal_to_float(delivery_total_pieces),
            "delivery_total_weight": decimal_to_float(delivery_total_weight),
            "delivery_total_pallets": decimal_to_float(delivery_total_pallets),
            "source_total_cartons": decimal_to_float(source_total_cartons),
            "source_total_weight": decimal_to_float(source_total_weight),
            "source_total_pallets": decimal_to_float(source_total_pallets),
            "source_total_loose_cartons": decimal_to_float(source_total_loose_cartons),
            "duplicate_delivery_fee_notes": duplicate_delivery_fee_notes,
            "delivery_validation_issue_count": len(self.delivery_validation_issues),
            "delivery_validation_issues": self.delivery_validation_issues,
        }


@dataclass
class AppliedInvoiceResult:
    invoice_name: str
    matched_row: int
    customer_order_no: str
    invoice_number: str
    loading_remark: str
    inbound_fee: Decimal
    multi_point_delivery_fee: Decimal
    loading_fee: Decimal

    def to_preview(self) -> Dict[str, Any]:
        return {
            "invoice_name": self.invoice_name,
            "matched_row": self.matched_row,
            "customer_order_no": self.customer_order_no,
            "invoice_number": self.invoice_number,
            "loading_remark": self.loading_remark,
            "inbound_fee": decimal_to_float(self.inbound_fee),
            "multi_point_delivery_fee": decimal_to_float(self.multi_point_delivery_fee),
            "loading_fee": decimal_to_float(self.loading_fee),
        }


@dataclass
class ReconcileOutput:
    output_path: Optional[Path]
    applied: List[AppliedInvoiceResult]
    errors: List[str]
    invoice_previews: List[Dict[str, Any]]
    master_preview_headers: List[str]
    master_preview_rows: List[Dict[str, Any]]


def parse_decimal_cell(value: object) -> Decimal:
    if value in (None, ""):
        return Decimal("0.00")
    return parse_amount(value)


def parse_decimal_value(value: object) -> Decimal:
    if value is None:
        return Decimal("0.00")
    if isinstance(value, float) and pd.isna(value):
        return Decimal("0.00")
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return Decimal("0.00")
    return parse_amount(text)


def normalize_invoice_transport_no(path: Path, invoice_number: str) -> str:
    for candidate in [invoice_number, path.stem, path.name]:
        text = str(candidate or "")
        match = re.search(r"(03FlEXH\d+)", text, flags=re.IGNORECASE)
        if match:
            return match.group(1)
    return path.stem


def load_delivery_note_rows(path: Path) -> Dict[int, Dict[int, object]]:
    workbook_path = make_excel_compatible_copy(path)
    return read_xlsx_rows(workbook_path)


def parse_delivery_note_files(
    delivery_note_paths: Sequence[Path],
) -> Dict[str, List[DeliveryNoteLine]]:
    grouped: Dict[str, List[DeliveryNoteLine]] = {}
    for path in delivery_note_paths:
        rows = load_delivery_note_rows(path)
        for row_idx in sorted(rows):
            if row_idx == 1:
                continue
            row = rows[row_idx]
            delivery_no = str(row.get(2) or "").strip()
            if not delivery_no:
                continue
            match = re.match(r"(.+)-(\d+)$", delivery_no)
            invoice_number = match.group(1) if match else delivery_no
            line = DeliveryNoteLine(
                delivery_no=delivery_no,
                invoice_number=invoice_number,
                delivery_time=str(row.get(3) or "").strip(),
                warehouse_so=str(row.get(7) or "").strip(),
                company_name=str(row.get(4) or "").strip(),
                area=str(row.get(5) or "").strip(),
                address=str(row.get(6) or "").strip(),
                total_pieces=parse_decimal_cell(row.get(11)),
                pallet_count=parse_decimal_cell(row.get(12)),
                billing_pallet_count=parse_decimal_cell(row.get(13)),
                carton_count=parse_decimal_cell(row.get(14)),
                gross_weight=parse_decimal_cell(row.get(15)),
            )
            grouped.setdefault(invoice_number, []).append(line)
    return grouped


def normalize_customer_order_no(value: str) -> str:
    match = re.search(r"(EXTR\d+)", str(value or "").upper())
    return match.group(1) if match else str(value or "").strip().upper()


def normalize_tan_no(value: str) -> str:
    text = str(value or "").strip().upper()
    match = re.search(r"TAN#?\s*(\d+)", text)
    if match:
        return f"TAN#{match.group(1)}"
    return text


def normalize_header_name(value: object) -> str:
    return str(value or "").strip().replace("\n", "").replace(" ", "").lower()


def row_values_as_text(row: pd.Series) -> List[str]:
    values: List[str] = []
    for value in row.tolist():
        if pd.isna(value):
            continue
        text = str(value).strip()
        if text and text.lower() != "nan":
            values.append(text)
    return values


def is_numeric_like(value: object) -> bool:
    if value is None or pd.isna(value):
        return False
    if isinstance(value, (int, float)):
        return True
    text = str(value).strip()
    return bool(re.fullmatch(r"\d+(?:\.\d+)?", text))


def find_tan_no_in_row(row: pd.Series) -> str:
    for text in row_values_as_text(row):
        stripped = text.strip()
        match = re.match(r"^TAN#?\s*(\d+)\b", stripped, flags=re.IGNORECASE)
        if match:
            return f"TAN#{match.group(1)}"
    return ""


def build_source_header_map(header_row: pd.Series) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for idx, value in enumerate(header_row.tolist()):
        key = normalize_header_name(value)
        if key:
            mapping[key] = idx
    return mapping


def pick_source_column(header_map: Dict[str, int], *candidates: str) -> Optional[int]:
    normalized_candidates = [normalize_header_name(item) for item in candidates]
    for candidate in normalized_candidates:
        if candidate in header_map:
            return header_map[candidate]
    return None


def is_source_detail_row(row: pd.Series) -> bool:
    first = "" if pd.isna(row.iloc[0]) else str(row.iloc[0]).strip()
    second = "" if len(row) < 2 or pd.isna(row.iloc[1]) else str(row.iloc[1]).strip()
    if is_numeric_like(row.iloc[0]) or first.isdigit():
        return True
    return not first and second.upper().startswith("SN-")


def build_source_note_text(row: pd.Series, tan_no: str) -> str:
    texts = row_values_as_text(row)
    if not texts:
        return ""
    cleaned: List[str] = []
    for text in texts:
        compact = re.sub(r"\s+", "", text.upper())
        if tan_no and normalize_tan_no(text) == tan_no:
            continue
        if compact == tan_no.replace(" ", ""):
            continue
        if re.fullmatch(r"\d+(?:\.\d+)?", text):
            continue
        cleaned.append(text)
    return "\n".join(cleaned).strip()


def parse_source_file(path: Path) -> tuple[str, List[SourceTanRecord]]:
    return cached_file_result(
        "invoice.parse_source_file",
        path,
        lambda: _parse_source_file_uncached(path),
    )


def _parse_source_file_uncached(path: Path) -> tuple[str, List[SourceTanRecord]]:
    customer_order_no = normalize_customer_order_no(path.stem)
    engine = "xlrd" if path.suffix.lower() == ".xls" else "openpyxl"
    workbook = pd.ExcelFile(path, engine=engine)
    sheet_name = workbook.sheet_names[0]
    df = workbook.parse(sheet_name, header=None)

    header_row_idx: Optional[int] = None
    for idx in range(len(df)):
        first_value = str(df.iloc[idx, 0] if 0 in df.columns else "").strip()
        second_value = str(df.iloc[idx, 1] if 1 in df.columns else "").strip()
        if first_value == "Item" and "SN" in second_value.upper():
            header_row_idx = idx
            break
    if header_row_idx is None:
        return customer_order_no, []

    header_map = build_source_header_map(df.iloc[header_row_idx])
    carton_col = pick_source_column(header_map, "箱数", "总箱数")
    gross_weight_col = pick_source_column(header_map, "毛重KG", "毛重")
    pallet_col = pick_source_column(header_map, "卡板数", "板数")
    if carton_col is None or gross_weight_col is None or pallet_col is None:
        return customer_order_no, []

    records: List[SourceTanRecord] = []
    pending_rows: List[pd.Series] = []
    pending_notes: List[str] = []

    def append_pending_record(tan_no: str) -> None:
        carton_count = sum((parse_decimal_value(item.iloc[carton_col]) for item in pending_rows), Decimal("0.00"))
        pallet_count = sum((parse_decimal_value(item.iloc[pallet_col]) for item in pending_rows), Decimal("0.00"))
        loose_carton_count = sum(
            (
                parse_decimal_value(item.iloc[carton_col])
                for item in pending_rows
                if parse_decimal_value(item.iloc[pallet_col]) <= Decimal("0.00")
            ),
            Decimal("0.00"),
        )
        gross_weight = sum((parse_decimal_value(item.iloc[gross_weight_col]) for item in pending_rows), Decimal("0.00"))
        records.append(
            SourceTanRecord(
                customer_order_no=customer_order_no,
                tan_no=tan_no,
                description="\n".join(item for item in pending_notes if item).strip(),
                carton_count=carton_count,
                pallet_count=pallet_count,
                loose_carton_count=loose_carton_count,
                gross_weight=gross_weight,
                row_count=len(pending_rows),
            )
        )

    for idx in range(header_row_idx + 1, len(df)):
        row = df.iloc[idx]
        row_texts = row_values_as_text(row)
        if not row_texts:
            continue

        tan_no = find_tan_no_in_row(row)
        if tan_no:
            tan_note = build_source_note_text(row, tan_no)
            if tan_note:
                pending_notes.append(tan_note)
            if pending_rows:
                append_pending_record(tan_no)
                pending_rows = []
                pending_notes = []
            continue

        if is_source_detail_row(row):
            pending_rows.append(row)
            continue

        if pending_rows:
            filtered_texts = [text for text in row_texts if not re.fullmatch(r"\d+(?:\.\d+)?", text)]
            note_text = "\n".join(filtered_texts).strip()
            if note_text:
                pending_notes.append(note_text)

    return customer_order_no, records


def parse_source_files(
    source_paths: Sequence[Path],
) -> Dict[str, List[SourceTanRecord]]:
    grouped: Dict[str, List[SourceTanRecord]] = {}
    for path in source_paths:
        customer_order_no, records = parse_source_file(path)
        grouped[customer_order_no] = records
    return grouped


def delivery_point_key(line: DeliveryNoteLine) -> str:
    tan_no = normalize_tan_no(line.warehouse_so)
    if tan_no:
        return tan_no
    return normalize_delivery_fee_note(f"{line.company_name}{line.address}")


def unique_delivery_point_count(delivery_note_lines: Sequence[DeliveryNoteLine]) -> int:
    keys = {delivery_point_key(line) for line in delivery_note_lines if delivery_point_key(line)}
    return len(keys) if keys else len(delivery_note_lines)


def build_delivery_validation_issues(
    invoice: InvoiceData,
    delivery_note_lines: Sequence[DeliveryNoteLine],
    source_tan_records: Sequence[SourceTanRecord],
) -> List[Dict[str, Any]]:
    issues: List[Dict[str, Any]] = []
    if not source_tan_records:
        issues.append(
            {
                "type": "missing_source_data",
                "title": "未导入对应真实数据源",
                "detail": "这张单页账单当前没有匹配到对应的真实数据源文件，暂时无法按 Tan# 校验件数、板数和毛重。",
            }
        )
        return issues

    if not delivery_note_lines:
        issues.append(
            {
                "type": "missing_delivery_note",
                "title": "未导入对应派送单",
                "detail": "这张单页账单当前没有匹配到派送单，暂时无法把真实数据源逐条比对到派送单。",
            }
        )
    else:
        source_map = {normalize_tan_no(item.tan_no): item for item in source_tan_records}
        delivery_map = {
            normalize_tan_no(item.warehouse_so): item
            for item in delivery_note_lines
            if item.warehouse_so
        }

        missing_in_delivery = sorted(set(source_map) - set(delivery_map))
        if missing_in_delivery:
            issues.append(
                {
                    "type": "source_tan_missing_in_delivery_note",
                    "title": "真实数据源中的 Tan# 未出现在派送单",
                    "detail": "以下 Tan# 在真实数据源里存在，但在派送单里没有找到对应仓单号SO：" + "、".join(missing_in_delivery[:10]),
                }
            )

        missing_in_source = sorted(set(delivery_map) - set(source_map))
        if missing_in_source:
            issues.append(
                {
                    "type": "delivery_tan_missing_in_source",
                    "title": "派送单中的 Tan# 未出现在真实数据源",
                    "detail": "以下仓单号SO在派送单里存在，但在真实数据源里没有找到对应 Tan#：" + "、".join(missing_in_source[:10]),
                }
            )

        mismatches: List[str] = []
        for tan_no in sorted(set(source_map) & set(delivery_map)):
            source = source_map[tan_no]
            delivery = delivery_map[tan_no]
            diff_parts: List[str] = []
            if source.carton_count != delivery.total_pieces:
                diff_parts.append(f"件数/箱数 {decimal_to_float(delivery.total_pieces):.2f}/{decimal_to_float(source.carton_count):.2f}")
            if source.pallet_count != delivery.pallet_count:
                diff_parts.append(f"板数 {decimal_to_float(delivery.pallet_count):.2f}/{decimal_to_float(source.pallet_count):.2f}")
            if source.gross_weight != delivery.gross_weight:
                diff_parts.append(f"毛重 {decimal_to_float(delivery.gross_weight):.2f}/{decimal_to_float(source.gross_weight):.2f}")
            if diff_parts:
                mismatches.append(f"{tan_no}（{'，'.join(diff_parts)}）")
        if mismatches:
            issues.append(
                {
                    "type": "delivery_source_value_mismatch",
                    "title": "派送单与真实数据源数值不一致",
                    "detail": "以下 Tan# 的派送单录入与真实数据源不一致：" + "；".join(mismatches[:10]),
                }
            )

        delivery_point_count = unique_delivery_point_count(delivery_note_lines)
        expected_delivery_fee = DELIVERY_UNIT * Decimal(delivery_point_count)
        if invoice.multi_point_delivery_fee < expected_delivery_fee:
            issues.append(
                {
                    "type": "delivery_fee_missing_by_point_count",
                    "title": "疑似少录派送费",
                    "detail": (
                        f"派送单当前有 {delivery_point_count} 个送货点，按 239.16/点应为 "
                        f"{decimal_to_float(expected_delivery_fee):.2f}，单页账单派送费为 "
                        f"{decimal_to_float(invoice.multi_point_delivery_fee):.2f}。请检查是否有送货点只录了装卸费。"
                    ),
                }
            )
        elif invoice.multi_point_delivery_fee > expected_delivery_fee:
            issues.append(
                {
                    "type": "delivery_fee_exceeds_point_count",
                    "title": "派送费金额超过派送单点数",
                    "detail": (
                        f"派送单当前有 {delivery_point_count} 个送货点，按 239.16/点应为 "
                        f"{decimal_to_float(expected_delivery_fee):.2f}，单页账单派送费为 "
                        f"{decimal_to_float(invoice.multi_point_delivery_fee):.2f}。请检查是否多录派送费。"
                    ),
                }
            )

        expected_loading_pallets = sum((item.pallet_count for item in source_tan_records), Decimal("0.00"))
        expected_loading_cartons = sum((item.loose_carton_count for item in source_tan_records), Decimal("0.00"))
        expected_loading_amount = quantized(expected_loading_pallets * PALLET_UNIT + expected_loading_cartons * CARTON_UNIT)
        actual_loading_pallets = Decimal(invoice.loading.pallet_count)
        actual_loading_cartons = Decimal(invoice.loading.carton_count)
        if (
            actual_loading_pallets != expected_loading_pallets
            or actual_loading_cartons != expected_loading_cartons
        ):
            issues.append(
                {
                    "type": "loading_split_mismatch_source",
                    "title": "装卸费板/箱数与真实数据源不一致",
                    "detail": (
                        f"单页账单装卸费当前拆为 {decimal_to_float(actual_loading_pallets):.2f} 板/"
                        f"{decimal_to_float(actual_loading_cartons):.2f} 箱；真实数据源按计费规则应为 "
                        f"{decimal_to_float(expected_loading_pallets):.2f} 板/"
                        f"{decimal_to_float(expected_loading_cartons):.2f} 散箱。请检查是否把快递费或错误箱数计入装卸费。"
                    ),
                }
            )

        suspected_express_rows = [
            f"第 {item.row_number} 行「{item.description or '-'}」金额为 {decimal_to_float(item.amount):.2f}"
            for item in invoice.fee_items
            if (
                "装卸费" in item.normalized_fee_name
                and item.amount == EXPRESS_UNIT
                and parse_loading_count_hint(item.description) == (Decimal("0.00"), Decimal("0.00"))
            )
        ]
        if invoice.loading.total_amount > expected_loading_amount and suspected_express_rows:
            issues.append(
                {
                    "type": "loading_fee_may_be_express",
                    "title": "装卸费 30 疑似快递费",
                    "detail": (
                        f"真实数据源应收装卸费为 {decimal_to_float(expected_loading_amount):.2f}，"
                        f"单页账单装卸费为 {decimal_to_float(invoice.loading.total_amount):.2f}。"
                        "以下装卸费金额刚好等于快递费单价 30，且说明中没有明确板/箱数量："
                        + "；".join(suspected_express_rows[:10])
                        + "。请确认是否把快递费录成了装卸费。"
                    ),
                }
            )

        delivery_fee_notes = {
            normalize_delivery_fee_note(item.description)
            for item in invoice.fee_items
            if "派送费" in item.normalized_fee_name and normalize_delivery_fee_note(item.description)
        }
        missing_delivery_fee_notes: List[str] = []
        for item in invoice.fee_items:
            if "装卸费" not in item.normalized_fee_name:
                continue
            normalized_note = normalize_delivery_fee_note(item.description)
            if normalized_note and normalized_note not in delivery_fee_notes:
                missing_delivery_fee_notes.append(f"第 {item.row_number} 行「{item.description}」")
        if missing_delivery_fee_notes:
            issues.append(
                {
                    "type": "loading_note_without_delivery_fee",
                    "title": "装卸费地点缺少对应派送费",
                    "detail": (
                        "以下装卸费说明出现了送货地点，但单页账单派送费里没有对应备注："
                        + "；".join(missing_delivery_fee_notes[:10])
                        + "。请检查是否少录派送费。"
                    ),
                }
            )

    duplicate_delivery_groups = find_duplicate_delivery_fee_notes(invoice.fee_items)
    if duplicate_delivery_groups:
        repeated = "；".join(
            f"{item['display_note']}（{item['count']}笔）"
            for item in duplicate_delivery_groups[:5]
        )
        source_total_cartons = sum((item.carton_count for item in source_tan_records), Decimal("0.00"))
        source_total_gross_weight = sum((item.gross_weight for item in source_tan_records), Decimal("0.00"))
        issues.append(
            {
                "type": "delivery_fee_note_duplicate",
                "title": "派送费说明疑似重复",
                "detail": f"这张账单的派送费说明里出现了重复送货备注：{repeated}。请结合真实数据源和派送单检查是否重复收费。真实数据源当前合计箱数 {decimal_to_float(source_total_cartons):.2f}，毛重 {decimal_to_float(source_total_gross_weight):.2f}。",
            }
        )

    return issues


def parse_invoice(path: Path) -> InvoiceData:
    price_reference = load_price_reference()
    rows = read_xlsx_rows(path)
    customer_order_no = find_value_beside_label(rows, ["客户订单号"])
    invoice_number = find_value_beside_label(rows, ["发票号码"])
    job_date_raw = find_value_beside_label(rows, ["作业时间"])
    truck_type_raw = find_value_beside_label(rows, ["车型", "车型："])
    origin_city = find_value_beside_label(rows, ["起运地"])
    destination_city = find_value_beside_label(rows, ["目的地"])

    if not customer_order_no:
        raise ValueError(f"{path.name} 未找到客户订单号")
    if not job_date_raw:
        raise ValueError(f"{path.name} 未找到作业时间")

    try:
        job_date = excel_serial_to_datetime(float(job_date_raw))
    except ValueError:
        job_date = datetime.fromisoformat(str(job_date_raw).strip())

    fee_items: List[InvoiceFeeItem] = []
    suspected_fee_name_errors: List[Dict[str, Any]] = []
    currencies: List[str] = []
    totals = {
        "freight": Decimal("0.00"),
        "seamless_fee": Decimal("0.00"),
        "registration_fee": Decimal("0.00"),
        "parking_fee": Decimal("0.00"),
        "multi_point_delivery_fee": Decimal("0.00"),
        "airport_fee": Decimal("0.00"),
        "container_rental_fee": Decimal("0.00"),
        "express_fee": Decimal("0.00"),
        "loading_total": Decimal("0.00"),
    }

    for row_idx in sorted(rows):
        row = rows[row_idx]
        fee_name = str(row.get(1) or "").strip()
        amount_raw = row.get(5)
        if not fee_name or amount_raw in (None, ""):
            continue

        description = str(row.get(3) or "").strip()
        amount = parse_amount(amount_raw)
        normalized_name, correction_reason = maybe_correct_fee_name(fee_name, description, amount)
        currency_match = re.search(r"[A-Za-z]{3}", str(amount_raw))
        currency = currency_match.group(0).upper() if currency_match else "HKD"
        currencies.append(currency)

        item = InvoiceFeeItem(
            row_number=row_idx,
            fee_name=fee_name,
            description=description,
            amount=amount,
            currency=currency,
            normalized_fee_name=normalized_name,
            corrected_fee_name=normalized_name if correction_reason else "",
            suspected_name_error=bool(correction_reason),
        )
        fee_items.append(item)

        if correction_reason:
            suspected_fee_name_errors.append(
                {
                    "row_number": row_idx,
                    "original_fee_name": fee_name,
                    "corrected_fee_name": normalized_name,
                    "description": description,
                    "amount": decimal_to_float(amount),
                    "reason": correction_reason,
                }
            )

        if "中港运费" in normalized_name:
            totals["freight"] += amount
        elif "无缝" in normalized_name:
            totals["seamless_fee"] += amount
        elif "租柜费" in normalized_name:
            totals["container_rental_fee"] += amount
        elif "登记费" in normalized_name:
            totals["registration_fee"] += amount
        elif "停车费" in normalized_name:
            totals["parking_fee"] += amount
        elif "装卸费" in normalized_name:
            totals["loading_total"] += amount
        elif "派送费" in normalized_name:
            totals["multi_point_delivery_fee"] += amount
        elif "机场附加费" in normalized_name or normalized_name == "机场费":
            totals["airport_fee"] += amount
        elif any(key in normalized_name for key in ["快递费/邮费", "快递费", "邮费"]):
            totals["express_fee"] += amount

    loading = split_loading(totals["loading_total"])
    currency = currencies[0] if currencies else "HKD"
    normalized_truck_type = normalize_truck_type(truck_type_raw)
    route_key = (
        normalize_route_city(origin_city),
        normalize_route_city(destination_city),
        normalize_price_truck_type(normalized_truck_type or truck_type_raw),
    )
    quoted_freight = price_reference.get(route_key)
    validation_issues = build_validation_issues(
        freight=quantized(totals["freight"]),
        seamless_fee=quantized(totals["seamless_fee"]),
        container_rental_fee=quantized(totals["container_rental_fee"]),
        multi_point_delivery_fee=quantized(totals["multi_point_delivery_fee"]),
        airport_fee=quantized(totals["airport_fee"]),
        express_fee=quantized(totals["express_fee"]),
        raw_truck_type=truck_type_raw,
        normalized_truck_type=normalized_truck_type,
        origin_city=origin_city,
        destination_city=destination_city,
        price_reference=price_reference,
    )
    validation_issues.extend(build_fee_line_validation_issues(fee_items))
    return InvoiceData(
        invoice_path=path,
        invoice_number=invoice_number or "",
        customer_order_no=customer_order_no,
        job_date=job_date,
        truck_type=normalized_truck_type,
        origin_city=origin_city,
        destination_city=destination_city,
        currency=currency,
        freight=quantized(totals["freight"]),
        seamless_fee=quantized(totals["seamless_fee"]),
        registration_fee=quantized(totals["registration_fee"]),
        parking_fee=quantized(totals["parking_fee"]),
        inbound_fee=quantized(totals["registration_fee"]),
        multi_point_delivery_fee=quantized(totals["multi_point_delivery_fee"]),
        airport_fee=quantized(totals["airport_fee"]),
        container_rental_fee=quantized(totals["container_rental_fee"]),
        express_fee=quantized(totals["express_fee"]),
        loading=loading,
        fee_items=fee_items,
        suspected_fee_name_errors=suspected_fee_name_errors,
        validation_issues=validation_issues,
        quoted_freight=quoted_freight,
        delivery_note_lines=[],
        source_tan_records=[],
        delivery_validation_issues=[],
    )


def build_header_map(ws: Worksheet) -> Dict[str, int]:
    result: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(MASTER_HEADER_ROW, col).value
        if header:
            result[str(header).strip()] = col
    return result


def find_target_row(ws: Worksheet, sli_col: int, sli_value: str) -> Optional[int]:
    for row in range(MASTER_DATA_START_ROW, ws.max_row + 1):
        if str(ws.cell(row, sli_col).value).strip() == sli_value:
            return row
    return None


def find_first_writable_row(ws: Worksheet, sli_col: int) -> int:
    for row in range(MASTER_DATA_START_ROW, ws.max_row + 1):
        if ws.cell(row, sli_col).value in (None, ""):
            return row
    return ws.max_row + 1


def find_style_source_row(ws: Worksheet, target_row: int, sli_col: int) -> int:
    for row in range(target_row - 1, MASTER_DATA_START_ROW - 1, -1):
        has_value = any(ws.cell(row, col).value not in (None, "") for col in range(1, ws.max_column + 1))
        has_style = any(ws.cell(row, col).has_style for col in range(1, ws.max_column + 1))
        if has_value or has_style:
            return row
    return MASTER_DATA_START_ROW


def clone_row_style(ws: Worksheet, source_row: int, target_row: int) -> None:
    if source_row == target_row:
        return

    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(source_row, col)
        target_cell = ws.cell(target_row, col)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)

    if source_row in ws.row_dimensions:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
        ws.row_dimensions[target_row].hidden = ws.row_dimensions[source_row].hidden


def set_default_fields(ws: Worksheet, row: int, header_map: Dict[str, int]) -> None:
    for header, default_value in FIXED_DEFAULTS.items():
        col = header_map.get(header)
        if col and ws.cell(row, col).value in (None, ""):
            ws.cell(row, col).value = default_value

    item_col = header_map.get("序号")
    if item_col and ws.cell(row, item_col).value in (None, ""):
        ws.cell(row, item_col).value = row - MASTER_DATA_START_ROW + 1


def update_master_row(ws: Worksheet, row: int, header_map: Dict[str, int], invoice: InvoiceData) -> None:
    def set_value(header: str, value: object) -> None:
        ws.cell(row, header_map[header]).value = value

    source_row = find_style_source_row(ws, row, header_map["SLI 跟踪单号"])
    clone_row_style(ws, source_row, row)
    set_default_fields(ws, row, header_map)
    set_value("年份", invoice.job_date.year)
    set_value("月份", invoice.job_date.month)
    set_value("SLI 跟踪单号", invoice.customer_order_no)
    if invoice.truck_type:
        set_value("车型", invoice.truck_type)
    set_value("装货时间", invoice.job_date)
    set_value("运费", decimal_to_float(invoice.freight))
    set_value("香港无缝费", decimal_to_float(invoice.seamless_fee))
    set_value("进仓费", decimal_to_float(invoice.registration_fee))
    set_value("停车费", decimal_to_float(invoice.parking_fee))
    set_value("多点送货费", decimal_to_float(invoice.multi_point_delivery_fee))
    set_value("机场费", decimal_to_float(invoice.airport_fee))
    set_value("租柜费", decimal_to_float(invoice.container_rental_fee))
    set_value("传真快递费", decimal_to_float(invoice.express_fee))
    set_value("币种", invoice.currency)

    loading_col = header_map["装卸费"]
    remark_col = header_map["备注板数"]
    ws.cell(row, remark_col).value = invoice.loading.remark
    if invoice.loading.exact_split:
        ws.cell(row, loading_col).value = f"=BP{row}*63.78"
    else:
        ws.cell(row, loading_col).value = decimal_to_float(invoice.loading.total_amount)

    ws.cell(row, header_map["应收合计-不含税"]).value = f"=SUM(AS{row}:BM{row})"
    ws.cell(row, header_map["应收合计-含税"]).value = f"=SUM(BN{row}:BN{row})"


def get_preview_headers(ws: Worksheet, limit: Optional[int] = None) -> List[str]:
    headers: List[str] = []
    max_col = limit or ws.max_column
    for col in range(1, max_col + 1):
        headers.append(str(ws.cell(MASTER_HEADER_ROW, col).value or f"COL_{col}"))
    return headers


def format_preview_value(header: str, value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if header in MONEY_COLUMNS and value is not None:
        try:
            return f"{float(value):.2f}"
        except (TypeError, ValueError):
            return value
    return value


def build_master_preview_rows(ws: Worksheet, headers: List[str]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for row_idx in range(MASTER_DATA_START_ROW, ws.max_row + 1):
        row_dict: Dict[str, Any] = {}
        has_value = False
        for offset, header in enumerate(headers, start=1):
            value = ws.cell(row_idx, offset).value
            row_dict[header] = format_preview_value(header, value)
            if value not in (None, ""):
                has_value = True
        if has_value:
            rows.append(row_dict)
    return rows


def reconcile(
    master_path: Path,
    invoice_paths: Sequence[Path],
    delivery_note_paths: Optional[Sequence[Path]] = None,
    source_paths: Optional[Sequence[Path]] = None,
    output_path: Optional[Path] = None,
) -> ReconcileOutput:
    workbook = openpyxl.load_workbook(master_path)
    ws = workbook[workbook.sheetnames[0]]
    header_map = build_header_map(ws)
    required_headers = [
        "年份",
        "月份",
        "SLI 跟踪单号",
        "车型",
        "装货时间",
        "运费",
        "香港无缝费",
        "进仓费",
        "停车费",
        "多点送货费",
        "机场费",
        "租柜费",
        "装卸费",
        "传真快递费",
        "币种",
        "应收合计-不含税",
        "应收合计-含税",
        "备注板数",
    ]
    missing = [item for item in required_headers if item not in header_map]
    if missing:
        raise ValueError(f"总表缺少必要列: {', '.join(missing)}")

    applied: List[AppliedInvoiceResult] = []
    errors: List[str] = []
    invoice_previews: List[Dict[str, Any]] = []
    sli_col = header_map["SLI 跟踪单号"]
    delivery_note_map = parse_delivery_note_files(delivery_note_paths or [])
    source_map = parse_source_files(source_paths or [])

    for invoice_path in invoice_paths:
        try:
            invoice = parse_invoice(invoice_path)
            transport_no = normalize_invoice_transport_no(invoice_path, invoice.invoice_number)
            delivery_note_lines = delivery_note_map.get(transport_no, [])
            invoice.delivery_note_lines = delivery_note_lines
            invoice.source_tan_records = source_map.get(normalize_customer_order_no(invoice.customer_order_no), [])
            invoice.delivery_validation_issues = build_delivery_validation_issues(
                invoice,
                delivery_note_lines,
                invoice.source_tan_records,
            )
            row = find_target_row(ws, sli_col, invoice.customer_order_no)
            if row is None:
                row = find_first_writable_row(ws, sli_col)

            update_master_row(ws, row, header_map, invoice)
            invoice_previews.append(
                {
                    **invoice.summary_preview(),
                    "matched_row": row,
                    "transport_no": transport_no,
                    "fee_items": [item.to_preview() for item in invoice.fee_items],
                }
            )
            applied.append(
                AppliedInvoiceResult(
                    invoice_name=invoice.invoice_path.name,
                    matched_row=row,
                    customer_order_no=invoice.customer_order_no,
                    invoice_number=invoice.invoice_number,
                    loading_remark=invoice.loading.remark,
                    inbound_fee=invoice.inbound_fee,
                    multi_point_delivery_fee=invoice.multi_point_delivery_fee,
                    loading_fee=invoice.loading.total_amount,
                )
            )
        except Exception as exc:  # noqa: BLE001
            errors.append(f"{invoice_path.name}: {exc}")

    headers = get_preview_headers(ws)
    preview_rows = build_master_preview_rows(ws, headers)
    if output_path:
        workbook.save(output_path)

    return ReconcileOutput(
        output_path=output_path,
        applied=applied,
        errors=errors,
        invoice_previews=invoice_previews,
        master_preview_headers=headers,
        master_preview_rows=preview_rows,
    )


def collect_invoice_paths(raw_inputs: Sequence[str]) -> List[Path]:
    paths: List[Path] = []
    for item in raw_inputs:
        path = Path(item)
        if path.is_dir():
            for pattern in ["*.xlsx", "*.xls"]:
                paths.extend(sorted(p for p in path.glob(pattern) if not p.name.startswith("~$")))
        else:
            paths.append(path)
    return paths


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="根据单页账单批量回填总账单")
    parser.add_argument("--master", required=True, help="总表 xlsx 路径")
    parser.add_argument("--invoices", nargs="+", required=True, help="单页账单文件或目录")
    parser.add_argument("--delivery-notes", nargs="*", default=[], help="派送单文件或目录")
    parser.add_argument("--sources", nargs="*", default=[], help="真实数据源文件或目录")
    parser.add_argument("--output", required=True, help="输出 xlsx 路径")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    result = reconcile(
        master_path=Path(args.master).expanduser().resolve(),
        invoice_paths=collect_invoice_paths(args.invoices),
        delivery_note_paths=collect_invoice_paths(args.delivery_notes),
        source_paths=collect_invoice_paths(args.sources),
        output_path=Path(args.output).expanduser().resolve(),
    )

    print(f"已输出: {result.output_path}")
    print(f"成功: {len(result.applied)}")
    for item in result.applied:
        print(
            "  - "
            f"{item.invoice_name} -> 第 {item.matched_row} 行, "
            f"SLI={item.customer_order_no}, "
            f"进仓费={item.inbound_fee}, "
            f"多点送货费={item.multi_point_delivery_fee}, "
            f"装卸费={item.loading_fee}, "
            f"备注={item.loading_remark}"
        )

    if result.errors:
        print(f"失败: {len(result.errors)}")
        for message in result.errors:
            print(f"  - {message}")
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
