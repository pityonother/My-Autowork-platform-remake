from __future__ import annotations

import json
import re
import shutil
import sqlite3
import subprocess
import uuid
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable, Optional, Sequence
from xml.etree import ElementTree as ET
from zipfile import ZipFile

from app.core.db import connect, run_migrations
from app.shared.lazy_imports import lazy_module
from app.shared.performance import cached_file_result
from app_paths import RUNTIME_DIR
from customs_reconciler import load_excel_file

pd = lazy_module("pandas")
xlrd = lazy_module("xlrd")


DB_PATH = RUNTIME_DIR / "finance_records.db"
SPECIAL_HANDLING_FEE = Decimal("70.00")

CATEGORY_FY = "fy_export"
CATEGORY_FT = "ft_export"
CATEGORY_OTHER = "other"

TASK_STATUS_OPTIONS = [
    ("submitted_to_finance", "提交给财务付款"),
    ("water_slip_received", "财务提供水单"),
    ("sent_out", "水单已发送出去"),
    ("invoice_received", "对方开票回来"),
    ("completed", "任务完成"),
]
TASK_STATUS_LABELS = dict(TASK_STATUS_OPTIONS)


@dataclass
class FinanceImportInput:
    payment_path: Path


@dataclass
class FinanceExportRow:
    record_id: int
    reimbursement_date: str
    forwarder_inv_no: str
    hawb_ref: str
    forwarder: str
    special_handling: Decimal
    amount_hkd: Decimal
    amount_note: str
    currency: str
    category: str
    remark: str
    batch_code: str


def get_connection() -> sqlite3.Connection:
    return connect(DB_PATH)


def migration_001_initial_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS finance_batches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_code TEXT NOT NULL UNIQUE,
            source_file_name TEXT NOT NULL,
            record_count INTEGER NOT NULL,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS finance_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id INTEGER NOT NULL,
            source_row_no INTEGER NOT NULL,
            expense_name TEXT NOT NULL,
            amount_rmb REAL NOT NULL,
            amount_hkd REAL NOT NULL,
            currency TEXT NOT NULL,
            reimbursement_date TEXT NOT NULL,
            remark TEXT,
            so_customer_raw TEXT,
            so_no TEXT,
            hawb_ref TEXT,
            forwarder TEXT,
            category TEXT NOT NULL,
            is_beryl INTEGER NOT NULL,
            is_bill_exportable INTEGER NOT NULL,
            invoice_required INTEGER NOT NULL,
            task_status TEXT NOT NULL,
            bill_exported_at TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(batch_id) REFERENCES finance_batches(id)
        );

        CREATE INDEX IF NOT EXISTS idx_finance_records_batch ON finance_records(batch_id);
        CREATE INDEX IF NOT EXISTS idx_finance_records_currency ON finance_records(currency);
        CREATE INDEX IF NOT EXISTS idx_finance_records_category ON finance_records(category);
        CREATE INDEX IF NOT EXISTS idx_finance_records_task_status ON finance_records(task_status);
        CREATE INDEX IF NOT EXISTS idx_finance_records_reimbursement_date ON finance_records(reimbursement_date);
        """
    )


MIGRATIONS = {
    1: migration_001_initial_schema,
}


def init_finance_db() -> None:
    with get_connection() as conn:
        run_migrations(conn, MIGRATIONS)
        conn.execute(
            """
            UPDATE finance_records
            SET is_bill_exportable = 0
            WHERE is_bill_exportable = 1
              AND (
                TRIM(COALESCE(so_no, '')) = ''
                OR TRIM(COALESCE(hawb_ref, '')) = ''
                OR TRIM(COALESCE(forwarder, '')) = ''
              )
            """
        )


def quantized(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def parse_decimal(value: object) -> Decimal:
    if value is None:
        return Decimal("0.00")
    if not isinstance(value, str) and pd.isna(value):
        return Decimal("0.00")
    text = str(value).strip().replace(",", "")
    if not text or text.lower() == "nan":
        return Decimal("0.00")
    try:
        return quantized(Decimal(text))
    except Exception:  # noqa: BLE001
        return Decimal("0.00")


def clean_text(value: object) -> str:
    if value is None:
        return ""
    if not isinstance(value, str) and pd.isna(value):
        return ""
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return ""
    return text


def parse_date_text(value: object) -> str:
    if value is None or (not isinstance(value, str) and pd.isna(value)):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, (int, float)) and 20000 <= float(value) <= 80000:
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date().isoformat()
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.replace(".", "", 1).isdigit():
            serial_value = float(stripped)
            if 20000 <= serial_value <= 80000:
                return (datetime(1899, 12, 30) + timedelta(days=serial_value)).date().isoformat()
    parsed = pd.to_datetime(value)
    if pd.isna(parsed):
        return ""
    return parsed.date().isoformat()


def normalize_party_text(value: object) -> str:
    text = clean_text(value)
    text = text.replace("（", "(").replace("）", ")")
    text = text.strip(" .。")
    return "".join(text.split()).lower()


def classify_remark(remark: str) -> str:
    normalized = normalize_party_text(remark)
    if "伟创力物联网科技" in normalized:
        return CATEGORY_OTHER
    if "伟创力电子设备" in normalized:
        return CATEGORY_FY
    if "福田伟创力代垫" in remark:
        return CATEGORY_FT
    if "伟创力出口代垫" in remark:
        return CATEGORY_FY
    return CATEGORY_OTHER


def is_beryl_remark(remark: str) -> bool:
    normalized = normalize_party_text(remark)
    return "beryl" in normalized or "beyrl" in normalized or "伟创力电子设备" in normalized


def split_so_customer(raw_value: str) -> tuple[str, str, str]:
    parts = [part.strip() for part in raw_value.split("\\") if part.strip()]
    if len(parts) >= 3:
        return parts[0], parts[1], parts[2]
    if len(parts) == 2:
        return parts[0], parts[1], ""
    if len(parts) == 1:
        return parts[0], "", ""
    return "", "", ""


PAYMENT_AMOUNT_RE = re.compile(r"(RMB|HKD)\s*([+-]?\d[\d,]*(?:\.\d+)?)", re.IGNORECASE)
PAYMENT_DESCRIPTION_DATE_RE = re.compile(
    r"(?:\d{4}[/-]\d{1,2}[/-]\d{1,2}|\d{1,2}[/-]\d{1,2}(?:[/-]\d{2,4})?)"
)


def parse_payment_amount_text(value: object) -> tuple[str, Decimal, Decimal]:
    text = clean_text(value)
    if not text:
        return "", Decimal("0.00"), Decimal("0.00")
    match = PAYMENT_AMOUNT_RE.search(text.replace(" ", ""))
    if not match:
        return "", Decimal("0.00"), Decimal("0.00")
    currency = match.group(1).upper()
    amount = parse_decimal(match.group(2))
    if currency == "RMB":
        return currency, amount, Decimal("0.00")
    return currency, Decimal("0.00"), amount


def parse_payment_description_date(value: str, fallback_date: str) -> str:
    text = clean_text(value)
    if not text or not PAYMENT_DESCRIPTION_DATE_RE.fullmatch(text):
        return ""
    if re.fullmatch(r"\d{1,2}[/-]\d{1,2}", text):
        year = int(fallback_date[:4]) if fallback_date else datetime.now().year
        month_text, day_text = re.split(r"[/-]", text)
        try:
            return date(year, int(month_text), int(day_text)).isoformat()
        except ValueError:
            return ""
    return parse_date_text(text)


def split_invoice_payment_description(raw_value: str, fallback_date: str) -> tuple[str, str, str, str, str]:
    parts = [part.strip() for part in raw_value.split("\\") if part.strip()]
    reimbursement_date = fallback_date
    if parts:
        parsed_date = parse_payment_description_date(parts[-1], fallback_date)
        if parsed_date:
            reimbursement_date = parsed_date
            parts = parts[:-1]
    business_raw = "\\".join(parts)
    so_no, hawb_ref, forwarder = split_so_customer(business_raw)
    return business_raw, so_no, hawb_ref, forwarder, reimbursement_date


def find_right_value(sheet: Any, row_index: int, label_col: int, max_col: int | None = None) -> object:
    max_col = max_col or sheet.max_column
    for col_index in range(label_col + 1, max_col + 1):
        value = sheet.cell(row_index, col_index).value
        if clean_text(value):
            return value
    return ""


class InvoicePaymentCell:
    def __init__(self, value: object = "") -> None:
        self.value = value


class InvoicePaymentXmlSheet:
    def __init__(self, values: dict[tuple[int, int], object]) -> None:
        self.values = values
        self.max_row = max((row for row, _ in values), default=0)
        self.max_column = max((col for _, col in values), default=0)

    def cell(self, row: int, column: int) -> InvoicePaymentCell:
        return InvoicePaymentCell(self.values.get((row, column), ""))


def xml_cell_column_index(cell_ref: str) -> int:
    result = 0
    for char in cell_ref:
        if not char.isalpha():
            break
        result = result * 26 + ord(char.upper()) - ord("A") + 1
    return result


def coerce_xml_cell_value(value: str) -> object:
    text = value.strip()
    if not text:
        return ""
    if re.fullmatch(r"[+-]?\d+", text):
        return int(text)
    if re.fullmatch(r"[+-]?\d+\.\d+", text):
        return float(text)
    return value


def read_xlsx_xml_cell_value(cell: ET.Element, shared_strings: Sequence[str], ns: dict[str, str]) -> object:
    cell_type = cell.attrib.get("t")
    if cell_type == "s":
        value_node = cell.find("a:v", ns)
        if value_node is None or value_node.text is None:
            return ""
        index = int(value_node.text)
        return shared_strings[index] if index < len(shared_strings) else ""
    if cell_type == "inlineStr":
        return "".join(text_node.text or "" for text_node in cell.findall(".//a:t", ns))
    value_node = cell.find("a:v", ns)
    if value_node is None or value_node.text is None:
        return ""
    return coerce_xml_cell_value(value_node.text)


def load_xlsx_first_sheet_xml(path: Path) -> InvoicePaymentXmlSheet:
    ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with ZipFile(path) as archive:
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            shared_root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in shared_root.findall("a:si", ns):
                shared_strings.append("".join(text_node.text or "" for text_node in item.findall(".//a:t", ns)))

        sheet_names = [
            name
            for name in archive.namelist()
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        ]
        if not sheet_names:
            return InvoicePaymentXmlSheet({})
        sheet_root = ET.fromstring(archive.read(sorted(sheet_names)[0]))

    values: dict[tuple[int, int], object] = {}
    for row_node in sheet_root.findall(".//a:sheetData/a:row", ns):
        row_index = int(row_node.attrib.get("r", "0") or 0)
        for cell_node in row_node.findall("a:c", ns):
            cell_ref = cell_node.attrib.get("r", "")
            col_index = xml_cell_column_index(cell_ref)
            if row_index <= 0 or col_index <= 0:
                continue
            values[(row_index, col_index)] = read_xlsx_xml_cell_value(cell_node, shared_strings, ns)
    return InvoicePaymentXmlSheet(values)


def load_invoice_payment_sheet(payment_path: Path) -> Any:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(payment_path, read_only=False, data_only=True)
        return workbook.worksheets[0]
    except Exception:  # noqa: BLE001
        return load_xlsx_first_sheet_xml(payment_path)


def find_invoice_payment_header(sheet: Any) -> tuple[int, int, int, int] | None:
    for row_index in range(1, sheet.max_row + 1):
        header_cols: dict[str, int] = {}
        for col_index in range(1, sheet.max_column + 1):
            text = normalize_party_text(sheet.cell(row_index, col_index).value)
            if "费用名称" in text:
                header_cols["expense"] = col_index
            elif "说明" in text:
                header_cols["description"] = col_index
            elif "金额" in text:
                header_cols["amount"] = col_index
        if {"expense", "description", "amount"}.issubset(header_cols):
            return row_index, header_cols["expense"], header_cols["description"], header_cols["amount"]
    return None


def read_invoice_payment_context(sheet: Any) -> tuple[str, str]:
    company_name = ""
    reimbursement_date = ""
    for row_index in range(1, sheet.max_row + 1):
        for col_index in range(1, sheet.max_column + 1):
            text = normalize_party_text(sheet.cell(row_index, col_index).value)
            if not company_name and "致" in text:
                company_name = clean_text(find_right_value(sheet, row_index, col_index))
            if not reimbursement_date and "作业时间" in text:
                reimbursement_date = parse_date_text(find_right_value(sheet, row_index, col_index))
        if company_name and reimbursement_date:
            return company_name, reimbursement_date
    return company_name, reimbursement_date


def parse_invoice_payment_rows(payment_path: Path) -> list[dict[str, Any]]:
    try:
        sheet = load_invoice_payment_sheet(payment_path)
    except Exception:  # noqa: BLE001
        return []

    header = find_invoice_payment_header(sheet)
    if header is None:
        return []

    header_row, expense_col, description_col, amount_col = header
    company_name, reimbursement_date = read_invoice_payment_context(sheet)
    if not company_name or not reimbursement_date:
        return []

    category = classify_remark(company_name)
    is_beryl = is_beryl_remark(company_name)
    parsed_rows: list[dict[str, Any]] = []
    for row_index in range(header_row + 1, sheet.max_row + 1):
        expense_name = clean_text(sheet.cell(row_index, expense_col).value)
        so_customer_raw = clean_text(sheet.cell(row_index, description_col).value)
        amount_text = clean_text(sheet.cell(row_index, amount_col).value)
        row_text = " ".join(
            clean_text(sheet.cell(row_index, col_index).value)
            for col_index in range(1, sheet.max_column + 1)
        )
        if not any([expense_name, so_customer_raw, amount_text]):
            if parsed_rows:
                break
            continue
        if "BANK NAME" in row_text.upper() or "PAGE " in row_text.upper():
            break
        if not expense_name or not so_customer_raw or not amount_text:
            continue

        currency, amount_rmb, amount_hkd = parse_payment_amount_text(amount_text)
        if not currency:
            continue
        business_raw, so_no, hawb_ref, forwarder, row_reimbursement_date = split_invoice_payment_description(
            so_customer_raw,
            reimbursement_date,
        )
        remark = " ".join(part for part in [company_name, expense_name, business_raw] if part)
        is_bill_exportable = int(bool(category == CATEGORY_FY and is_beryl and so_no and hawb_ref and forwarder))
        parsed_rows.append(
            {
                "source_row_no": row_index,
                "expense_name": expense_name,
                "amount_rmb": float(amount_rmb),
                "amount_hkd": float(amount_hkd),
                "currency": currency,
                "reimbursement_date": row_reimbursement_date,
                "remark": remark,
                "so_customer_raw": business_raw,
                "so_no": so_no,
                "hawb_ref": hawb_ref,
                "forwarder": forwarder,
                "category": category,
                "is_beryl": int(is_beryl),
                "is_bill_exportable": is_bill_exportable,
                "invoice_required": int(currency == "RMB"),
                "task_status": "submitted_to_finance",
                "allow_existing_update": True,
            }
        )
    return parsed_rows


def parse_legacy_payment_rows(payment_path: Path) -> list[dict[str, Any]]:
    excel_file, _ = load_excel_file(payment_path)
    df = excel_file.parse(excel_file.sheet_names[0])
    parsed_rows: list[dict[str, Any]] = []

    for idx, row in df.iterrows():
        if int(float(row.get("取消", 0) or 0)) == 1:
            continue
        expense_name = clean_text(row.get("支出名称"))
        reimbursement_date = parse_date_text(row.get("报销日期"))
        remark = clean_text(row.get("备注"))
        so_customer_raw = clean_text(row.get("SO/客户"))
        currency = clean_text(row.get("币种")).upper()
        amount_rmb = parse_decimal(row.get("RMB金额"))
        amount_hkd = parse_decimal(row.get("HKD金额"))
        if not expense_name or not reimbursement_date:
            continue

        category = classify_remark(remark)
        is_beryl = is_beryl_remark(remark)
        so_no, hawb_ref, forwarder = split_so_customer(so_customer_raw)
        is_bill_exportable = int(bool(category == CATEGORY_FY and is_beryl and so_no and hawb_ref and forwarder))
        parsed_rows.append(
            {
                "source_row_no": idx + 2,
                "expense_name": expense_name,
                "amount_rmb": float(amount_rmb),
                "amount_hkd": float(amount_hkd),
                "currency": currency,
                "reimbursement_date": reimbursement_date,
                "remark": remark,
                "so_customer_raw": so_customer_raw,
                "so_no": so_no,
                "hawb_ref": hawb_ref,
                "forwarder": forwarder,
                "category": category,
                "is_beryl": int(is_beryl),
                "is_bill_exportable": is_bill_exportable,
                "invoice_required": int(currency == "RMB"),
                "task_status": "submitted_to_finance",
            }
        )
    return parsed_rows


def parse_finance_payment_rows(payment_path: Path) -> list[dict[str, Any]]:
    return cached_file_result(
        "finance.parse_finance_payment_rows",
        payment_path,
        lambda: _parse_finance_payment_rows_uncached(payment_path),
    )


def _parse_finance_payment_rows_uncached(payment_path: Path) -> list[dict[str, Any]]:
    try:
        rows = parse_legacy_payment_rows(payment_path)
    except Exception:  # noqa: BLE001
        rows = []
    if rows:
        return rows
    return parse_invoice_payment_rows(payment_path)


def find_duplicate_finance_record(conn: sqlite3.Connection, row: dict[str, Any]) -> sqlite3.Row | None:
    amount = finance_business_amount(row)
    candidates = conn.execute(
        """
        SELECT *
        FROM finance_records
        WHERE reimbursement_date = ?
          AND currency = ?
          AND ABS(CASE WHEN currency = 'HKD' THEN amount_hkd ELSE amount_rmb END - ?) < 0.01
        ORDER BY (TRIM(COALESCE(so_customer_raw, '')) <> '') DESC, id DESC
        """,
        (row["reimbursement_date"], row["currency"], amount),
    ).fetchall()
    target_remark = normalize_business_text(row["remark"])
    for candidate in candidates:
        if normalize_business_text(candidate["remark"] or "") == target_remark:
            return candidate
        if row.get("allow_existing_update") and finance_record_business_fields_match(candidate, row):
            return candidate
    if row.get("allow_existing_update"):
        candidates = conn.execute(
            """
            SELECT *
            FROM finance_records
            WHERE currency = ?
              AND ABS(CASE WHEN currency = 'HKD' THEN amount_hkd ELSE amount_rmb END - ?) < 0.01
            ORDER BY id DESC
            """,
            (row["currency"], amount),
        ).fetchall()
        for candidate in candidates:
            if finance_record_business_fields_match(candidate, row):
                return candidate
            if normalize_business_text(candidate["remark"] or "") == target_remark:
                return candidate
    return None


def update_existing_finance_record(conn: sqlite3.Connection, existing: sqlite3.Row, row: dict[str, Any], updated_at: str) -> bool:
    conn.execute(
        """
        UPDATE finance_records
        SET source_row_no = ?,
            expense_name = ?,
            amount_rmb = ?,
            amount_hkd = ?,
            currency = ?,
            reimbursement_date = ?,
            remark = ?,
            so_customer_raw = ?,
            so_no = ?,
            hawb_ref = ?,
            forwarder = ?,
            category = ?,
            is_beryl = ?,
            is_bill_exportable = ?,
            invoice_required = ?,
            updated_at = ?
        WHERE id = ?
        """,
        (
            row["source_row_no"],
            row["expense_name"],
            row["amount_rmb"],
            row["amount_hkd"],
            row["currency"],
            row["reimbursement_date"],
            row["remark"],
            row["so_customer_raw"],
            row["so_no"],
            row["hawb_ref"],
            row["forwarder"],
            row["category"],
            row["is_beryl"],
            row["is_bill_exportable"],
            row["invoice_required"],
            updated_at,
            existing["id"],
        ),
    )
    return True


def merge_duplicate_finance_record(conn: sqlite3.Connection, existing: sqlite3.Row, row: dict[str, Any], updated_at: str) -> bool:
    if row.get("allow_existing_update"):
        return update_existing_finance_record(conn, existing, row, updated_at)
    if not row["so_customer_raw"]:
        return False
    if (existing["so_customer_raw"] or "").strip():
        return False
    conn.execute(
        """
        UPDATE finance_records
        SET so_customer_raw = ?,
            so_no = ?,
            hawb_ref = ?,
            forwarder = ?,
            is_bill_exportable = ?,
            updated_at = ?
        WHERE id = ?
        """,
        (
            row["so_customer_raw"],
            row["so_no"],
            row["hawb_ref"],
            row["forwarder"],
            row["is_bill_exportable"],
            updated_at,
            existing["id"],
        ),
    )
    return True


def import_finance_batch(import_input: FinanceImportInput) -> dict[str, Any]:
    init_finance_db()

    now = datetime.now()
    batch_code = f"FIN-{now.strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"
    rows_to_insert = parse_finance_payment_rows(import_input.payment_path)

    with get_connection() as conn:
        cursor = conn.execute(
            """
            INSERT INTO finance_batches (batch_code, source_file_name, record_count, created_at)
            VALUES (?, ?, ?, ?)
            """,
            (
                batch_code,
                import_input.payment_path.name,
                len(rows_to_insert),
                now.isoformat(timespec="seconds"),
            ),
        )
        batch_id = cursor.lastrowid
        duplicate_count = 0
        merged_count = 0
        now_text = now.isoformat(timespec="seconds")
        insert_rows: list[tuple[Any, ...]] = []
        for row in rows_to_insert:
            existing = find_duplicate_finance_record(conn, row)
            if existing is not None:
                duplicate_count += 1
                if merge_duplicate_finance_record(conn, existing, row, now_text):
                    merged_count += 1
                continue
            insert_rows.append(
                (
                    batch_id,
                    row["source_row_no"],
                    row["expense_name"],
                    row["amount_rmb"],
                    row["amount_hkd"],
                    row["currency"],
                    row["reimbursement_date"],
                    row["remark"],
                    row["so_customer_raw"],
                    row["so_no"],
                    row["hawb_ref"],
                    row["forwarder"],
                    row["category"],
                    row["is_beryl"],
                    row["is_bill_exportable"],
                    row["invoice_required"],
                    row["task_status"],
                    now_text,
                    now_text,
                )
            )
        if insert_rows:
            conn.executemany(
                """
                INSERT INTO finance_records (
                    batch_id, source_row_no, expense_name, amount_rmb, amount_hkd, currency,
                    reimbursement_date, remark, so_customer_raw, so_no, hawb_ref, forwarder,
                    category, is_beryl, is_bill_exportable, invoice_required, task_status,
                    bill_exported_at, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, NULL, ?, ?)
                """,
                insert_rows,
            )
        inserted_count = len(insert_rows)
        conn.execute(
            "UPDATE finance_batches SET record_count = ? WHERE id = ?",
            (inserted_count, batch_id),
        )
    detail = get_finance_batch_detail(int(batch_id))
    detail["import_stats"] = {
        "parsed_count": len(rows_to_insert),
        "inserted_count": inserted_count,
        "duplicate_count": duplicate_count,
        "merged_count": merged_count,
        "skipped_count": duplicate_count - merged_count,
    }
    return detail


def list_finance_batches() -> list[dict[str, Any]]:
    init_finance_db()
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT
                b.*,
                SUM(CASE WHEN r.is_bill_exportable = 1 THEN 1 ELSE 0 END) AS exportable_count,
                SUM(CASE WHEN r.invoice_required = 1 THEN 1 ELSE 0 END) AS invoice_required_count
            FROM finance_batches b
            LEFT JOIN finance_records r ON r.batch_id = b.id
            GROUP BY b.id
            ORDER BY b.created_at DESC
            LIMIT 10000
            """
        ).fetchall()
    return [dict(row) for row in rows]


def list_finance_records(
    *,
    batch_id: int | None = None,
    currency: str | None = None,
    category: str | None = None,
    task_status: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    only_exportable: bool = False,
    only_invoice_required: bool = False,
) -> list[dict[str, Any]]:
    init_finance_db()
    query = """
        SELECT r.*, b.batch_code, b.source_file_name
        FROM finance_records r
        JOIN finance_batches b ON b.id = r.batch_id
        WHERE 1=1
    """
    params: list[Any] = []
    if batch_id:
        query += " AND r.batch_id = ?"
        params.append(batch_id)
    if currency:
        query += " AND r.currency = ?"
        params.append(currency.upper())
    if category:
        query += " AND r.category = ?"
        params.append(category)
    if task_status:
        query += " AND r.task_status = ?"
        params.append(task_status)
    if date_from:
        query += " AND r.reimbursement_date >= ?"
        params.append(date_from)
    if date_to:
        query += " AND r.reimbursement_date <= ?"
        params.append(date_to)
    if only_exportable:
        query += " AND r.is_bill_exportable = 1"
    if only_invoice_required:
        query += " AND r.invoice_required = 1"
    query += " ORDER BY r.reimbursement_date DESC, r.id DESC LIMIT 10000"

    with get_connection() as conn:
        rows = conn.execute(query, params).fetchall()
    return [format_finance_record(row) for row in rows]


def get_finance_batch_detail(batch_id: int) -> dict[str, Any]:
    init_finance_db()
    with get_connection() as conn:
        batch = conn.execute("SELECT * FROM finance_batches WHERE id = ?", (batch_id,)).fetchone()
        if batch is None:
            raise ValueError("未找到财务记录批次。")
    records = list_finance_records(batch_id=batch_id)
    return {"batch": dict(batch), "records": records}


def format_finance_record(row: sqlite3.Row) -> dict[str, Any]:
    category = row["category"]
    category_label = {
        CATEGORY_FY: "福永伟创力代垫",
        CATEGORY_FT: "福田伟创力代垫",
        CATEGORY_OTHER: "其他代支",
    }.get(category, category)
    task_status = row["task_status"]
    return {
        "id": row["id"],
        "batch_id": row["batch_id"],
        "batch_code": row["batch_code"],
        "source_file_name": row["source_file_name"],
        "source_row_no": row["source_row_no"],
        "expense_name": row["expense_name"],
        "amount_rmb": float(row["amount_rmb"] or 0),
        "amount_hkd": float(row["amount_hkd"] or 0),
        "currency": row["currency"],
        "reimbursement_date": row["reimbursement_date"],
        "remark": row["remark"] or "",
        "so_customer_raw": row["so_customer_raw"] or "",
        "so_no": row["so_no"] or "",
        "hawb_ref": row["hawb_ref"] or "",
        "forwarder": row["forwarder"] or "",
        "category": category,
        "category_label": category_label,
        "is_beryl": bool(row["is_beryl"]),
        "is_bill_exportable": bool(row["is_bill_exportable"]),
        "is_pending_export": not bool(row["bill_exported_at"]),
        "invoice_required": bool(row["invoice_required"]),
        "task_status": task_status,
        "task_status_label": TASK_STATUS_LABELS.get(task_status, task_status),
        "bill_exported_at": row["bill_exported_at"] or "",
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def update_finance_task_status(record_id: int, task_status: str) -> None:
    if task_status not in TASK_STATUS_LABELS:
        raise ValueError("不支持的任务状态。")
    init_finance_db()
    with get_connection() as conn:
        conn.execute(
            """
            UPDATE finance_records
            SET task_status = ?, updated_at = ?
            WHERE id = ?
            """,
            (task_status, datetime.now().isoformat(timespec="seconds"), record_id),
        )


def update_finance_export_status(record_id: int, is_exported: bool) -> None:
    init_finance_db()
    now = datetime.now().isoformat(timespec="seconds")
    exported_at = now if is_exported else None
    with get_connection() as conn:
        conn.execute(
            """
            UPDATE finance_records
            SET bill_exported_at = ?, updated_at = ?
            WHERE id = ?
            """,
            (exported_at, now, record_id),
        )


def read_outbound_existing_keys(bill_path: Path) -> set[tuple[str, str, str, str, str]]:
    keys: set[tuple[str, str, str, str, str]] = set()
    try:
        book = xlrd.open_workbook(str(bill_path), formatting_info=False)
    except Exception:
        return keys
    try:
        sheet = book.sheet_by_name("OUTBOUND")
    except Exception:
        if not book.sheet_names():
            return keys
        sheet = book.sheet_by_index(0)

    for row_index in range(sheet.nrows):
        row_text = " ".join(str(sheet.cell_value(row_index, col_index) or "") for col_index in range(min(sheet.ncols, 10)))
        if "TOTAL" in row_text.upper():
            break
        if sheet.ncols <= 8:
            continue
        forwarder_inv_no = clean_text(sheet.cell_value(row_index, 1) if sheet.ncols > 1 else "")
        hawb_ref = clean_text(sheet.cell_value(row_index, 3) if sheet.ncols > 3 else "")
        forwarder = clean_text(sheet.cell_value(row_index, 4) if sheet.ncols > 4 else "")
        if not (forwarder_inv_no or hawb_ref or forwarder):
            continue
        special_handling = parse_decimal(sheet.cell_value(row_index, 7) if sheet.ncols > 7 else 0)
        amount_hkd = parse_decimal(sheet.cell_value(row_index, 8) if sheet.ncols > 8 else 0)
        if amount_hkd == Decimal("0.00"):
            continue
        keys.add(
            (
                finance_export_text_key(forwarder_inv_no),
                finance_export_text_key(hawb_ref),
                finance_export_text_key(forwarder),
                finance_export_amount_key(special_handling),
                finance_export_amount_key(amount_hkd),
            )
        )
    return keys


def build_finance_export_rows(
    *,
    exchange_rate: Decimal,
    batch_id: int | None = None,
    currency: str | None = None,
    task_status: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    only_exportable: bool = False,
    only_invoice_required: bool = False,
) -> list[FinanceExportRow]:
    records = list_finance_records(
        batch_id=batch_id,
        currency=currency,
        category=CATEGORY_FY,
        task_status=task_status,
        date_from=date_from,
        date_to=date_to,
        only_exportable=only_exportable,
        only_invoice_required=only_invoice_required,
    )
    export_rows: list[FinanceExportRow] = []
    seen_keys: set[tuple[str, str, str, str]] = set()
    for row in records:
        if not row["is_bill_exportable"]:
            continue
        if not row["so_no"] or not row["hawb_ref"] or not row["forwarder"]:
            continue
        key = finance_business_key(row)
        if key in seen_keys:
            continue
        seen_keys.add(key)
        amount_note = ""
        if row["currency"].upper() == "HKD":
            amount_hkd = quantized(Decimal(str(row["amount_hkd"])))
        else:
            amount_rmb = quantized(Decimal(str(row["amount_rmb"])))
            if exchange_rate <= 0:
                raise ValueError("汇率必须大于 0。")
            amount_hkd = quantized(amount_rmb / exchange_rate)
            amount_note = f"{amount_rmb:.2f}RMB/{exchange_rate:.2f}={amount_hkd:.2f}HKD"
        export_rows.append(
            FinanceExportRow(
                record_id=row["id"],
                reimbursement_date=row["reimbursement_date"],
                forwarder_inv_no=row["so_no"],
                hawb_ref=row["hawb_ref"],
                forwarder=row["forwarder"],
                special_handling=SPECIAL_HANDLING_FEE,
                amount_hkd=amount_hkd,
                amount_note=amount_note,
                currency=row["currency"],
                category=row["category"],
                remark=row["remark"],
                batch_code=row["batch_code"],
            )
        )
    return export_rows


def summarize_finance_records(records: Sequence[dict[str, Any]]) -> dict[str, int]:
    return {
        "total_count": len(records),
        "exportable_count": sum(1 for row in records if row["is_bill_exportable"]),
        "invoice_required_count": sum(1 for row in records if row["invoice_required"]),
        "pending_export_count": sum(1 for row in records if row["is_bill_exportable"] and row["is_pending_export"]),
        "exported_count": sum(1 for row in records if bool(row["bill_exported_at"])),
    }


def mark_finance_records_exported(record_ids: Iterable[int]) -> None:
    ids = list(record_ids)
    if not ids:
        return
    init_finance_db()
    now = datetime.now().isoformat(timespec="seconds")
    placeholders = ",".join("?" for _ in ids)
    with get_connection() as conn:
        conn.execute(
            f"""
            UPDATE finance_records
            SET bill_exported_at = ?, updated_at = ?
            WHERE id IN ({placeholders})
            """,
            [now, now, *ids],
        )


def export_finance_outbound_bill(
    template_path: Path,
    rows: Sequence[FinanceExportRow],
    output_path: Path,
) -> list[int]:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template_path, output_path)

    existing_keys = read_outbound_existing_keys(template_path)
    rows_to_write: list[FinanceExportRow] = []
    processed_existing_ids: list[int] = []
    for row in rows:
        key = finance_export_row_key(row)
        if key in existing_keys:
            processed_existing_ids.append(row.record_id)
            continue
        rows_to_write.append(row)
        existing_keys.add(key)

    payload = [
        {
            "record_id": row.record_id,
            "inv_date": row.reimbursement_date,
            "forwarder_inv_no": row.forwarder_inv_no,
            "hawb_ref": row.hawb_ref,
            "forwarder": row.forwarder,
            "special_handling": float(row.special_handling),
            "amount_hkd": float(row.amount_hkd),
            "amount_note": row.amount_note,
        }
        for row in rows_to_write
    ]

    temp_json = output_path.with_suffix(".finance.json")
    temp_ps1 = output_path.with_suffix(".finance.ps1")
    temp_result = output_path.with_suffix(".finance-result.json")
    output_literal = str(output_path).replace("'", "''")
    json_literal = str(temp_json).replace("'", "''")
    result_literal = str(temp_result).replace("'", "''")
    powershell = f"""
$ErrorActionPreference = 'Stop'
$template = '{output_literal}'
$jsonPath = '{json_literal}'
$resultPath = '{result_literal}'
$rows = Get-Content -LiteralPath $jsonPath -Raw | ConvertFrom-Json
$excel = $null
$wb = $null
$ws = $null
$processedIds = New-Object System.Collections.Generic.List[int]
$closeWithSave = $false
try {{
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
$wb = $excel.Workbooks.Open($template)
$ws = $wb.Worksheets.Item(1)
$startRow = 5
$usedLastRow = $ws.UsedRange.Row + $ws.UsedRange.Rows.Count - 1
$insertBeforeRow = 0
$totalFormulaStartRow = 0

for ($rowIndex = $startRow; $rowIndex -le $usedLastRow; $rowIndex++) {{
    $rowText = ''
    for ($colIndex = 1; $colIndex -le 10; $colIndex++) {{
        $rowText += ([string]$ws.Cells.Item($rowIndex, $colIndex).Text)
    }}
    if ($rowText -match 'TOTAL') {{
        $insertBeforeRow = $rowIndex
        $totalFormula = [string]$ws.Cells.Item($rowIndex, 9).Formula
        if ($totalFormula -match '\\$?I\\$?(\\d+):\\$?I\\$?\\d+') {{
            $totalFormulaStartRow = [int]$Matches[1]
        }}
        break
    }}
}}
if ($insertBeforeRow -eq 0) {{
    $insertBeforeRow = $usedLastRow + 1
}}
if ($totalFormulaStartRow -le 0) {{
    $totalFormulaStartRow = $startRow + 1
}}

$lastDataRow = $startRow - 1
$blankRows = New-Object System.Collections.Queue
for ($rowIndex = $startRow; $rowIndex -lt $insertBeforeRow; $rowIndex++) {{
    $rowHasAnyValue = $false
    for ($colIndex = 1; $colIndex -le 10; $colIndex++) {{
        if ([string]$ws.Cells.Item($rowIndex, $colIndex).Text) {{
            $rowHasAnyValue = $true
            break
        }}
    }}
    if (-not $rowHasAnyValue) {{
        $blankRows.Enqueue($rowIndex)
        continue
    }}

    $b = [string]$ws.Cells.Item($rowIndex, 2).Text
    $d = [string]$ws.Cells.Item($rowIndex, 4).Text
    $e = [string]$ws.Cells.Item($rowIndex, 5).Text
    $iValue = $ws.Cells.Item($rowIndex, 9).Value2
    if ($b -or $d -or $e -or $iValue) {{
        $lastDataRow = $rowIndex
    }}
}}
$sampleRow = if ($lastDataRow -ge $startRow) {{ $lastDataRow }} else {{ $startRow }}

foreach ($row in @($rows)) {{
    if ($blankRows.Count -gt 0) {{
        $targetRow = [int]$blankRows.Dequeue()
    }} else {{
        $targetRow = $insertBeforeRow
        if ($targetRow -le $usedLastRow) {{
            $ws.Rows.Item($targetRow).Insert() | Out-Null
            $usedLastRow++
            $insertBeforeRow++
        }} else {{
            $insertBeforeRow++
        }}
        $ws.Range("A$sampleRow:J$sampleRow").Copy() | Out-Null
        $ws.Range("A$targetRow:J$targetRow").PasteSpecial(-4122) | Out-Null
    }}
    $ws.Cells.Item($targetRow, 1).Value = [datetime]::Parse($row.inv_date)
    $ws.Cells.Item($targetRow, 2).NumberFormat = '@'
    $ws.Cells.Item($targetRow, 4).NumberFormat = '@'
    $ws.Cells.Item($targetRow, 5).NumberFormat = '@'
    $ws.Cells.Item($targetRow, 2).Value = [string]$row.forwarder_inv_no
    $ws.Cells.Item($targetRow, 3).Value = ''
    $ws.Cells.Item($targetRow, 4).Value = [string]$row.hawb_ref
    $ws.Cells.Item($targetRow, 5).Value = [string]$row.forwarder
    $ws.Cells.Item($targetRow, 6).Value = ''
    $ws.Cells.Item($targetRow, 7).Value = ''
    $ws.Cells.Item($targetRow, 8).Value2 = [double]$row.special_handling
    $ws.Cells.Item($targetRow, 9).Value2 = [double]$row.amount_hkd
    $ws.Cells.Item($targetRow, 10).Value = ''
    try {{ $ws.Cells.Item($targetRow, 9).ClearComments() | Out-Null }} catch {{ }}
    if ([string]$row.amount_note) {{
        $ws.Cells.Item($targetRow, 9).AddComment([string]$row.amount_note) | Out-Null
    }}
    $processedIds.Add([int]$row.record_id) | Out-Null
}}

$totalRow = $insertBeforeRow
if ($totalRow -ge $totalFormulaStartRow) {{
    $formulaEndRow = $totalRow - 1
    if ($formulaEndRow -ge $totalFormulaStartRow) {{
        $ws.Cells.Item($totalRow, 9).Formula = "=SUM(I$($totalFormulaStartRow):I$($formulaEndRow))"
    }} else {{
        $ws.Cells.Item($totalRow, 9).Formula = "=0"
    }}
}}

$wb.Save()
$closeWithSave = $true
}} finally {{
    if ($null -ne $wb) {{
        try {{ $wb.Close($closeWithSave) }} catch {{ }}
    }}
    if ($null -ne $excel) {{
        try {{ $excel.Quit() }} catch {{ }}
    }}
    if ($null -ne $ws) {{
        try {{ [System.Runtime.Interopservices.Marshal]::ReleaseComObject($ws) | Out-Null }} catch {{ }}
    }}
    if ($null -ne $wb) {{
        try {{ [System.Runtime.Interopservices.Marshal]::ReleaseComObject($wb) | Out-Null }} catch {{ }}
    }}
    if ($null -ne $excel) {{
        try {{ [System.Runtime.Interopservices.Marshal]::ReleaseComObject($excel) | Out-Null }} catch {{ }}
    }}
}}
[PSCustomObject]@{{ processed_ids = @($processedIds) }} | ConvertTo-Json -Depth 4 | Set-Content -LiteralPath $resultPath -Encoding UTF8
"""
    try:
        temp_json.write_text(json.dumps(payload, ensure_ascii=True), encoding="utf-8")
        temp_ps1.write_text(powershell, encoding="utf-8")
        subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-File", str(temp_ps1)],
            check=True,
            capture_output=True,
            text=True,
        )
        if temp_result.exists():
            result = json.loads(temp_result.read_text(encoding="utf-8-sig"))
            return [*processed_existing_ids, *[int(record_id) for record_id in result.get("processed_ids", [])]]
        return processed_existing_ids
    finally:
        temp_json.unlink(missing_ok=True)
        temp_ps1.unlink(missing_ok=True)
        temp_result.unlink(missing_ok=True)


from app.modules.finance.parsers import parse_exchange_rate, parse_payment_amount_text, split_invoice_payment_description
from app.modules.finance.rules import (
    finance_business_amount,
    finance_business_key,
    finance_export_amount_key,
    finance_export_row_key,
    finance_export_text_key,
    finance_record_business_fields_match,
    normalize_business_text,
)
