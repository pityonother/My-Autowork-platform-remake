from __future__ import annotations

import asyncio
import json
import re
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.modules.booking.body_validation import (
    _parse_per_box_expression,
    build_body_validation_preview,
    build_corrected_body_validation_workbook,
)
from app.modules.booking import sil_fuca_delivery as delivery_module
from app.modules.booking.sil_fuca_delivery import (
    SilFucaDeliveryClient,
    SilFucaDeliveryRecord,
    SilFucaDeliveryResponse,
)
from app.modules.booking.schemas import BookingPreview
from app.modules.booking import routes as booking_routes
from booking_web_app import app as booking_app


def _workbook_bytes(rows: list[list[object]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Booking Form"
    for column_index, title in enumerate(
        [
            "Line",
            "Case number",
            "PO No.",
            "PN",
            "Part Description",
            "Quantity",
            "Unit",
            "Cartons",
            "N.Wt",
            "G.Wt",
            "CBM",
            "Pallet",
            "Invoice No.",
            "production date",
            "Invoice Date",
            "Made In",
            "Batch No.",
            "ASN",
            "Supplier card number",
            "Supplier delivery note number",
            "Tray Type",
            "brand",
            "LEDBinCode",
            "Min package",
            "Standard quantity per box",
            "IPPC",
            "Remark",
        ],
        start=1,
    ):
        ws.cell(8, column_index).value = title
    for offset, row in enumerate(rows, start=9):
        for column_index, value in enumerate(row, start=1):
            ws.cell(offset, column_index).value = value
    ws.cell(9 + len(rows), 1).value = "Total"
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _set_number_format(content: bytes, cell_address: str, number_format: str) -> bytes:
    wb = load_workbook(BytesIO(content))
    ws = wb[wb.sheetnames[0]]
    ws[cell_address].number_format = number_format
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _valid_row(*, made_date: object = "2026-01-12", per_box: object = 4000) -> list[object]:
    return [
        1,
        "CASE1",
        "C33C-25040701-0001",
        "ABC123",
        "IC",
        4000,
        "PCS",
        1,
        1,
        2,
        0.1,
        0,
        "INV1",
        made_date,
        "",
        "CN",
        "B1",
        "",
        "",
        "123456789012345",
        "",
        "BRAND",
        "BIN",
        1000,
        per_box,
        "",
        "",
    ]


def _sil_fuca_row(*, po: str, pn: str, quantity: object) -> list[object]:
    row = _valid_row()
    row[3 - 1] = po
    row[4 - 1] = pn
    row[6 - 1] = quantity
    return row


def _delivery_record(
    *,
    po: str,
    pn: str = "1010300202002T01",
    delivery_quantity: object = "20000",
    delivery_date: str = "2026-07-03",
) -> SilFucaDeliveryRecord:
    return SilFucaDeliveryRecord(
        po=po,
        product_code=pn,
        delivery_quantity=Decimal(str(delivery_quantity)),
        delivery_date=date.fromisoformat(delivery_date),
        delivery_list_no="2026062700003-0000000023",
    )


class _FakeSilFucaDeliveryClient:
    def __init__(
        self,
        *,
        new_records: dict[tuple[str, str], tuple[SilFucaDeliveryRecord, ...]] | None = None,
        new_errors: dict[tuple[str, str], tuple[str, ...]] | None = None,
        all_records: tuple[SilFucaDeliveryRecord, ...] = (),
    ) -> None:
        self.new_records = new_records or {}
        self.new_errors = new_errors or {}
        self.all_records = all_records
        self.queries = []
        self.all_call_count = 0

    def get_delivery_list_new(self, query):
        self.queries.append(query)
        key = (query.po, query.pn)
        records = self.new_records.get(key, ())
        return SilFucaDeliveryResponse(
            success=bool(records),
            records=records,
            errors=self.new_errors.get(key, ()),
        )

    def get_all_delivery_list(self):
        self.all_call_count += 1
        return self.all_records


class _CountingSilFucaDeliveryClient(SilFucaDeliveryClient):
    def __init__(self, *, raw_all_payload=None, error: str = "") -> None:
        super().__init__(list_all_url="http://local.test/all")
        self.raw_all_payload = raw_all_payload if raw_all_payload is not None else []
        self.error = error
        self.all_call_count = 0

    def _get_json(self, url: str):  # noqa: ANN001
        self.all_call_count += 1
        if self.error:
            raise RuntimeError(self.error)
        return self.raw_all_payload


def _reset_shared_delivery_cache() -> None:
    delivery_module._ALL_DELIVERY_LIST_CACHE.records = ()
    delivery_module._ALL_DELIVERY_LIST_CACHE.updated_at = None
    delivery_module._ALL_DELIVERY_LIST_CACHE.error = ""


def _delivery_api_payload(*, po: str = "T33U-26040025-0002", pn: str = "1010300202002T01") -> dict[str, object]:
    return {
        "po": po,
        "product_code": pn,
        "delivery_quantity": "20,000",
        "delivery_date": "2026-07-03T00:00:00",
        "delivery_list_no": "2026062700003-0000000023",
        "allocation_status": "未使用",
    }


def test_booking_body_validation_flags_static_body_issues() -> None:
    content = _workbook_bytes(
        [
            [
                1,
                "",
                "BADPO",
                "PN-1",
                "IC",
                0,
                "PCS",
                "1.5",
                0,
                0,
                "10X10",
                "NA",
                "INV #1",
                "bad-date",
                "",
                "MARS",
                "",
                "",
                "",
                "BAD",
                "WOODEN PALLET",
                "NA",
                "",
                10,
                3,
                "",
                "",
            ]
        ]
    )

    preview = build_body_validation_preview(content, filename="bad.xlsx")
    issue_fields = {issue.field_code for issue in preview.issues}

    assert preview.row_count == 1
    assert {
        "case_number",
        "PO_No",
        "Customer_Part_No",
        "Quantity",
        "Pkgs",
        "CBM",
        "Pallet",
        "Invoice_No",
        "madeDate",
        "Made_In",
        "packing",
        "brand",
        "per_box",
    }.issubset(issue_fields)
    assert "Pallet" in preview.rows[0].issue_fields


def test_booking_body_validation_applies_safe_static_fixes() -> None:
    content = _workbook_bytes(
        [
            [
                1,
                "CASE1",
                "C33C-25040701-0001",
                "ABC123",
                "IC",
                10,
                "",
                1,
                1,
                2,
                0,
                "NA",
                "INV #1",
                "202510",
                "",
                "CN",
                "",
                "",
                "",
                "123456789012345",
                "",
                "NA",
                "",
                "",
                "",
                "",
                "",
            ]
        ]
    )

    preview = build_body_validation_preview(content, filename="fix.xlsx", apply_fixes=True)
    row = preview.rows[0]

    assert row.values["unit"] == "PCS"
    assert row.values["Pallet"] == "0"
    assert row.values["Invoice_No"] == "INV1"
    assert row.values["Batch_No"] == "0"
    assert row.values["brand"] == "无"
    assert row.values["LEDBinCode"] == "无"
    assert row.values["min_package"] == "10"
    assert row.values["per_box"] == "10"
    assert preview.issue_count == 0
    assert preview.fix_count == 8


def test_booking_body_validation_keeps_manual_review_na_fields_unfixed() -> None:
    row = _valid_row(per_box=4000)
    row[8 - 1] = "NA"  # Cartons / Pkgs
    row[24 - 1] = "NA"  # Min package
    content = _workbook_bytes([row])

    preview = build_body_validation_preview(content, filename="manual-na.xlsx", apply_fixes=True)
    fixed_row = preview.rows[0]

    assert fixed_row.values["Pkgs"] == "NA"
    assert fixed_row.values["min_package"] == "NA"
    assert "Pkgs" in fixed_row.issue_fields
    assert "min_package" in fixed_row.issue_fields


def test_booking_body_validation_fills_empty_case_number_with_zero() -> None:
    row = _valid_row()
    row[2 - 1] = ""
    content = _workbook_bytes([row])

    preview = build_body_validation_preview(content, filename="case-number.xlsx", apply_fixes=True)

    assert preview.rows[0].values["case_number"] == "0"
    assert preview.rows[0].correction_kind_for("case_number") == "default_zero"
    assert "case_number" in preview.rows[0].fixed_fields
    assert preview.issue_count == 0

    corrected = build_corrected_body_validation_workbook(content, filename="case-number.xlsx")
    wb = load_workbook(BytesIO(corrected))
    assert wb[wb.sheetnames[0]]["B9"].value == "0"


def test_booking_body_validation_keeps_na_case_number_as_supplier_value() -> None:
    row = _valid_row()
    row[2 - 1] = "NA"
    content = _workbook_bytes([row])

    preview = build_body_validation_preview(content, filename="case-number-na.xlsx", apply_fixes=True)

    assert preview.rows[0].values["case_number"] == "NA"
    assert "case_number" not in preview.rows[0].fixed_fields
    assert preview.issue_count == 0


def test_booking_body_validation_keeps_na_case_number_when_other_fields_need_review() -> None:
    row = _valid_row()
    row[2 - 1] = "NA"
    row[8 - 1] = ""
    row[9 - 1] = ""
    row[10 - 1] = ""
    content = _workbook_bytes([row])

    preview = build_body_validation_preview(content, filename="case-number-na-with-missing-fields.xlsx", apply_fixes=True)

    assert preview.rows[0].values["case_number"] == "NA"
    assert "case_number" not in preview.rows[0].fixed_fields
    assert not any(issue.field_code == "case_number" for issue in preview.issues)


def test_booking_body_validation_normalizes_static_text_and_numeric_values() -> None:
    row = _valid_row()
    row[23 - 1] = "BIN"  # Keep LEDBinCode valid after changing nearby fields.
    row[24 - 1] = 1000
    row[25 - 1] = 4000
    row[3 - 1] = "C33C250407010001"
    row[4 - 1] = "AB-123 / 45"
    row[6 - 1] = "1,000 pcs"
    row[8 - 1] = "2+3"
    row[13 - 1] = "INV-001/02"
    row[16 - 1] = "TW"
    content = _workbook_bytes([row])

    preview = build_body_validation_preview(content, filename="normalize.xlsx", apply_fixes=True)
    fixed_row = preview.rows[0]

    assert fixed_row.values["PO_No"] == "C33C-25040701-0001"
    assert fixed_row.values["Customer_Part_No"] == "AB12345"
    assert fixed_row.values["Quantity"] == "1000"
    assert fixed_row.values["Pkgs"] == "5"
    assert fixed_row.values["Invoice_No"] == "INV00102"
    assert fixed_row.values["Made_In"] == "TW,CN"
    assert preview.issue_count == 0


def test_booking_body_validation_averages_zero_weight_by_case_number() -> None:
    first = _valid_row()
    first[2 - 1] = "CASE-A"
    first[9 - 1] = 0
    first[10 - 1] = 0
    second = _valid_row()
    second[1 - 1] = 2
    second[2 - 1] = "CASE-A"
    second[9 - 1] = 6
    second[10 - 1] = 8
    content = _workbook_bytes([first, second])

    preview = build_body_validation_preview(content, filename="weights.xlsx", apply_fixes=True)
    first_row, second_row = preview.rows

    assert first_row.values["FJZ"] == "3"
    assert second_row.values["FJZ"] == "3"
    assert first_row.values["G_Wt"] == "4"
    assert second_row.values["G_Wt"] == "4"
    assert first_row.correction_kind_for("FJZ") == "weight_average_by_case"
    assert second_row.correction_kind_for("G_Wt") == "weight_average_by_case"
    assert preview.issue_count == 0


def test_booking_body_validation_averages_missing_weight_with_previous_na_case_line() -> None:
    first = _valid_row(per_box=4000)
    first[2 - 1] = "NA"
    first[9 - 1] = 14.2
    first[10 - 1] = 14.8
    second = _valid_row(per_box="")
    second[1 - 1] = 2
    second[2 - 1] = "NA"
    second[8 - 1] = ""
    second[9 - 1] = ""
    second[10 - 1] = ""
    second[14 - 1] = ""
    second[17 - 1] = ""
    second[24 - 1] = ""
    second[25 - 1] = ""
    content = _workbook_bytes([first, second])

    preview = build_body_validation_preview(content, filename="weights-na.xlsx", apply_fixes=True)
    first_row, second_row = preview.rows

    assert first_row.values["FJZ"] == "7.1"
    assert second_row.values["FJZ"] == "7.1"
    assert first_row.values["G_Wt"] == "7.4"
    assert second_row.values["G_Wt"] == "7.4"
    assert first_row.correction_kind_for("FJZ") == "weight_average_previous_line"
    assert second_row.correction_kind_for("G_Wt") == "weight_average_previous_line"
    assert second_row.values["per_box"] == ""
    assert second_row.values["Batch_No"] == ""


def test_booking_body_validation_unmerges_weight_cells_when_exporting_previous_line_average() -> None:
    first = _valid_row(per_box=4000)
    first[2 - 1] = "NA"
    first[9 - 1] = 14.2
    first[10 - 1] = 14.8
    second = _valid_row(per_box="")
    second[1 - 1] = 2
    second[2 - 1] = "NA"
    second[8 - 1] = ""
    second[9 - 1] = ""
    second[10 - 1] = ""
    second[14 - 1] = ""
    second[17 - 1] = ""
    second[24 - 1] = ""
    second[25 - 1] = ""
    content = _workbook_bytes([first, second])
    wb = load_workbook(BytesIO(content))
    ws = wb[wb.sheetnames[0]]
    ws.merge_cells("I9:I10")
    ws.merge_cells("J9:J10")
    ws.merge_cells("Q9:Q10")
    buffer = BytesIO()
    wb.save(buffer)

    corrected = build_corrected_body_validation_workbook(buffer.getvalue(), filename="merged.xlsx")
    fixed_wb = load_workbook(BytesIO(corrected))
    fixed_ws = fixed_wb[fixed_wb.sheetnames[0]]

    assert "I9:I10" not in {str(item) for item in fixed_ws.merged_cells.ranges}
    assert "J9:J10" not in {str(item) for item in fixed_ws.merged_cells.ranges}
    assert fixed_ws["I9"].value == "7.1"
    assert fixed_ws["I10"].value == "7.1"
    assert fixed_ws["J9"].value == "7.4"
    assert fixed_ws["J10"].value == "7.4"
    assert fixed_ws["Q9"].value != "0"


def test_per_box_expression_parser_matches_supplier_examples() -> None:
    assert _parse_per_box_expression("2K+2K") == "4000"
    assert _parse_per_box_expression("6K*5CARTON+5K+3K*3CARTON+1K*2CARTON") == "46000"


def test_booking_body_validation_applies_example_corrections() -> None:
    content = _workbook_bytes(
        [
            _valid_row(made_date="2026-01-12 AND 2026-01-26", per_box="2K+2K"),
            _valid_row(per_box="6K*5carton+5k+3k*3carton+1k*2carton"),
        ]
    )

    preview = build_body_validation_preview(content, filename="supplier.xlsx", apply_fixes=True)
    first_row = preview.rows[0]
    second_row = preview.rows[1]

    assert first_row.source_values["madeDate"] == "2026-01-12AND2026-01-26"
    assert first_row.values["madeDate"] == "2026-01-12"
    assert first_row.correction_options_for("madeDate") == ("2026-01-12", "2026-01-26")
    assert first_row.correction_kind_for("madeDate") == "date_choice"
    assert first_row.values["per_box"] == "4000"
    assert second_row.values["per_box"] == "46000"
    assert "madeDate" in first_row.source_issue_fields
    assert "per_box" in first_row.source_issue_fields
    assert "per_box" in second_row.source_issue_fields
    assert preview.source_issue_count == 3
    assert preview.issue_count == 0


def test_booking_body_validation_fills_missing_per_box_from_quantity_and_cartons() -> None:
    row = _valid_row(per_box="")
    row[6 - 1] = 6000
    row[8 - 1] = 2
    row[24 - 1] = 1500
    content = _workbook_bytes([row])

    preview = build_body_validation_preview(content, filename="per-box-missing.xlsx", apply_fixes=True)

    assert preview.rows[0].values["per_box"] == "3000"
    assert preview.rows[0].correction_kind_for("per_box") == "per_box_from_quantity_cartons"
    assert preview.issue_count == 0


def test_booking_body_validation_fills_missing_per_box_from_min_package_when_cartons_is_zero() -> None:
    row = _valid_row(per_box="")
    row[6 - 1] = 15000
    row[8 - 1] = 0
    row[24 - 1] = 3000
    content = _workbook_bytes([row])

    preview = build_body_validation_preview(content, filename="per-box-from-min-package.xlsx", apply_fixes=True)

    assert preview.rows[0].values["per_box"] == "3000"
    assert preview.rows[0].correction_kind_for("per_box") == "per_box_from_min_package"
    assert not any(issue.field_code == "per_box" for issue in preview.issues)
    assert any(issue.field_code == "Pkgs" for issue in preview.issues)


def test_booking_body_validation_fills_missing_per_box_even_when_case_number_is_na() -> None:
    row = _valid_row(per_box="")
    row[2 - 1] = "NA"
    row[6 - 1] = 6000
    row[8 - 1] = 2
    row[24 - 1] = 1500
    content = _workbook_bytes([row])

    preview = build_body_validation_preview(content, filename="per-box-na-case.xlsx", apply_fixes=True)

    assert preview.rows[0].values["per_box"] == "3000"
    assert preview.rows[0].correction_kind_for("per_box") == "per_box_from_quantity_cartons"
    assert preview.issue_count == 0


def test_booking_body_validation_fills_missing_per_box_from_min_package_when_na_case_cartons_is_zero() -> None:
    row = _valid_row(per_box="")
    row[2 - 1] = "NA"
    row[6 - 1] = 10000
    row[8 - 1] = 0
    row[24 - 1] = 2500
    content = _workbook_bytes([row])

    preview = build_body_validation_preview(content, filename="per-box-na-from-min-package.xlsx", apply_fixes=True)

    assert preview.rows[0].values["case_number"] == "NA"
    assert preview.rows[0].values["per_box"] == "2500"
    assert preview.rows[0].correction_kind_for("per_box") == "per_box_from_min_package"
    assert "case_number" not in preview.rows[0].fixed_fields
    assert not any(issue.field_code == "per_box" for issue in preview.issues)
    assert any(issue.field_code == "Pkgs" for issue in preview.issues)


def test_booking_body_validation_strips_week_suffix_from_production_date() -> None:
    content = _workbook_bytes([_valid_row(made_date="2614M")])

    preview = build_body_validation_preview(content, filename="week-suffix.xlsx", apply_fixes=True)

    assert preview.rows[0].values["madeDate"] == "2614"
    assert preview.rows[0].correction_kind_for("madeDate") == "date_week_normalize"
    assert preview.issue_count == 0


def test_booking_body_validation_flags_excel_date_display_format() -> None:
    content = _workbook_bytes([_valid_row(made_date=datetime(2026, 3, 23))])
    content = _set_number_format(content, "N9", "mm/dd/yy")

    preview = build_body_validation_preview(content, filename="date-format.xlsx")

    date_issue = next(issue for issue in preview.issues if issue.field_code == "madeDate")
    assert date_issue.correction_kind == "date_format"
    assert "mm/dd/yy" in date_issue.message
    assert "2026/3/23" in date_issue.message
    assert "2026-03-23" in date_issue.message

    fixed_preview = build_body_validation_preview(content, filename="date-format.xlsx", apply_fixes=True)
    assert "madeDate" in fixed_preview.rows[0].fixed_fields
    assert fixed_preview.rows[0].correction_kind_for("madeDate") == "date_format"
    assert not fixed_preview.issues


def test_booking_body_validation_sil_fuca_delivery_accepts_successful_record() -> None:
    pn = "1010300202002T01"
    content = _workbook_bytes(
        [
            _sil_fuca_row(po="T33U-26040025-0002", pn=pn, quantity=12000),
            _sil_fuca_row(po="T33U-26040025-0002", pn=pn, quantity=8000),
        ]
    )
    client = _FakeSilFucaDeliveryClient(
        new_records={
            ("T33U-26040025-0002", pn): (
                _delivery_record(po="T33U-26040025-0002", pn=pn, delivery_quantity=20000),
            )
        }
    )

    preview = build_body_validation_preview(
        content,
        filename="sil.xlsx",
        apply_fixes=True,
        enable_dynamic_checks=True,
        sil_fuca_delivery_client=client,
        query_date=date(2026, 6, 29),
    )

    assert len(client.queries) == 1
    assert client.queries[0].qty == Decimal("20000")
    assert preview.issue_count == 0
    assert preview.fix_count == 0


def test_booking_body_validation_delivery_checks_use_shared_po_prefixes() -> None:
    pn = "1010176925000"
    content = _workbook_bytes([_valid_row(per_box=18000)])
    wb = load_workbook(BytesIO(content))
    ws = wb[wb.sheetnames[0]]
    ws["C9"] = "E33K-26030121-0009"
    ws["D9"] = pn
    ws["F9"] = 18000
    buffer = BytesIO()
    wb.save(buffer)
    client = _FakeSilFucaDeliveryClient(
        new_records={
            ("E33K-26030121-0009", pn): (
                _delivery_record(po="E33K-26030121-0009", pn=pn, delivery_quantity=48000),
            )
        }
    )

    preview = build_body_validation_preview(
        buffer.getvalue(),
        filename="e33k.xlsx",
        apply_fixes=True,
        enable_dynamic_checks=True,
        sil_fuca_delivery_client=client,
        query_date=date(2026, 6, 29),
    )

    assert len(client.queries) == 1
    assert client.queries[0].po == "E33K-26030121-0009"
    assert preview.rows[0].delivery_match_status == "ok"
    assert not any(issue.correction_kind.startswith("sil_fuca_delivery") for issue in preview.issues)


def test_sil_fuca_delivery_record_parses_api_date() -> None:
    record = SilFucaDeliveryRecord.from_api(
        {
            "purchase_order_type": "T33U",
            "purchase_order_no": "26040025",
            "purchase_order_seq": "0002",
            "product_code": "1010300202002T01",
            "delivery_quantity": 20000.0,
            "delivery_date": "2026-07-03T00:00:00",
        }
    )

    assert record.po == "T33U-26040025-0002"
    assert record.delivery_quantity == Decimal("20000.0")
    assert record.delivery_date == date(2026, 7, 3)


def test_sil_fuca_all_delivery_list_cache_persists_between_clients(monkeypatch, tmp_path) -> None:
    cache_file = tmp_path / "sil_fuca_all_delivery_list.json"
    monkeypatch.setenv("SIL_FUCA_DELIVERY_LIST_CACHE_FILE", str(cache_file))
    _reset_shared_delivery_cache()
    first_client = _CountingSilFucaDeliveryClient(raw_all_payload=[_delivery_api_payload()])

    first_records = first_client.get_all_delivery_list()

    assert first_client.all_call_count == 1
    assert cache_file.is_file()
    assert first_records[0].po == "T33U-26040025-0002"

    _reset_shared_delivery_cache()
    second_client = _CountingSilFucaDeliveryClient(error="should not call api")
    second_records = second_client.get_all_delivery_list()

    assert second_client.all_call_count == 0
    assert second_records[0].delivery_quantity == Decimal("20000")


def test_sil_fuca_all_delivery_list_uses_old_cache_when_refresh_lock_active(monkeypatch, tmp_path) -> None:
    cache_file = tmp_path / "sil_fuca_all_delivery_list.json"
    monkeypatch.setenv("SIL_FUCA_DELIVERY_LIST_CACHE_FILE", str(cache_file))
    _reset_shared_delivery_cache()
    seed_client = _CountingSilFucaDeliveryClient(raw_all_payload=[_delivery_api_payload()])
    seed_client.get_all_delivery_list()
    _reset_shared_delivery_cache()
    lock_path = cache_file.with_suffix(cache_file.suffix + ".lock")
    lock_path.write_text("refreshing", encoding="utf-8")
    refresh_client = _CountingSilFucaDeliveryClient(error="should not call api while locked")

    records = refresh_client.get_all_delivery_list(force_refresh=True)
    status = delivery_module.get_all_delivery_list_cache_status()

    assert refresh_client.all_call_count == 0
    assert records[0].po == "T33U-26040025-0002"
    assert status.state == "refreshing"
    assert status.record_count == 1


def test_sil_fuca_all_delivery_list_refresh_failure_preserves_persistent_cache(monkeypatch, tmp_path) -> None:
    cache_file = tmp_path / "sil_fuca_all_delivery_list.json"
    monkeypatch.setenv("SIL_FUCA_DELIVERY_LIST_CACHE_FILE", str(cache_file))
    _reset_shared_delivery_cache()
    seed_client = _CountingSilFucaDeliveryClient(raw_all_payload=[_delivery_api_payload()])
    seed_client.get_all_delivery_list()
    _reset_shared_delivery_cache()
    failing_client = _CountingSilFucaDeliveryClient(error="network down")

    records = failing_client.get_all_delivery_list(force_refresh=True)
    status = delivery_module.get_all_delivery_list_cache_status()

    assert failing_client.all_call_count == 1
    assert records[0].po == "T33U-26040025-0002"
    assert status.state == "error"
    assert "network down" in status.error
    assert "network down" in json.loads(cache_file.read_text(encoding="utf-8"))["error"]


def test_sil_fuca_scheduled_refresh_refreshes_missing_cache(monkeypatch, tmp_path) -> None:
    cache_file = tmp_path / "sil_fuca_all_delivery_list.json"
    monkeypatch.setenv("SIL_FUCA_DELIVERY_LIST_CACHE_FILE", str(cache_file))
    _reset_shared_delivery_cache()
    client = _CountingSilFucaDeliveryClient(raw_all_payload=[_delivery_api_payload()])

    refreshed = delivery_module.refresh_all_delivery_list_if_needed(client_factory=lambda: client)

    assert refreshed is True
    assert client.all_call_count == 1
    assert cache_file.is_file()


def test_sil_fuca_scheduled_refresh_skips_fresh_cache(monkeypatch, tmp_path) -> None:
    cache_file = tmp_path / "sil_fuca_all_delivery_list.json"
    monkeypatch.setenv("SIL_FUCA_DELIVERY_LIST_CACHE_FILE", str(cache_file))
    _reset_shared_delivery_cache()
    seed_client = _CountingSilFucaDeliveryClient(raw_all_payload=[_delivery_api_payload()])
    seed_client.get_all_delivery_list()
    failing_client = _CountingSilFucaDeliveryClient(error="should not call api")

    refreshed = delivery_module.refresh_all_delivery_list_if_needed(client_factory=lambda: failing_client)

    assert refreshed is False
    assert failing_client.all_call_count == 0


def test_sil_fuca_background_refresh_task_can_start_and_stop(monkeypatch) -> None:
    monkeypatch.setenv("SIL_FUCA_DELIVERY_LIST_AUTO_REFRESH", "1")
    monkeypatch.setenv("SIL_FUCA_DELIVERY_LIST_REFRESH_START_DELAY_SECONDS", "3600")

    async def run_task_check() -> None:
        await delivery_module.stop_delivery_list_background_refresh()
        delivery_module.start_delivery_list_background_refresh()
        assert delivery_module._BACKGROUND_REFRESH_TASK is not None
        assert not delivery_module._BACKGROUND_REFRESH_TASK.done()
        await delivery_module.stop_delivery_list_background_refresh()
        assert delivery_module._BACKGROUND_REFRESH_TASK is None

    asyncio.run(run_task_check())


def test_booking_body_validation_sil_fuca_delivery_suggests_unique_po_sequence() -> None:
    pn = "1010300202002T01"
    content = _workbook_bytes([_sil_fuca_row(po="T33U-26040025-0001", pn=pn, quantity=20000)])
    client = _FakeSilFucaDeliveryClient(
        new_errors={
            ("T33U-26040025-0001", pn): (
                "T33U-26040025-0001 1010300202002T01匹配不上周期交货清单，请检查是否有上传",
            )
        },
        all_records=(
            _delivery_record(po="T33U-26040025-0002", pn=pn, delivery_quantity=20000),
        ),
    )

    preview = build_body_validation_preview(
        content,
        filename="sil.xlsx",
        apply_fixes=True,
        enable_dynamic_checks=True,
        sil_fuca_delivery_client=client,
        query_date=date(2026, 6, 29),
    )

    row = preview.rows[0]
    assert client.all_call_count == 1
    assert row.values["PO_No"] == "T33U-26040025-0002"
    assert row.correction_kind_for("PO_No") == "sil_fuca_delivery_po"
    assert "PO_No" in row.fixed_fields
    assert "PO_No" in row.source_issue_fields
    assert preview.issue_count == 0
    assert preview.source_issue_count == 1

    corrected = build_corrected_body_validation_workbook(
        content,
        filename="sil.xlsx",
        enable_dynamic_checks=True,
        sil_fuca_delivery_client=client,
        query_date=date(2026, 6, 29),
    )
    wb = load_workbook(BytesIO(corrected))
    assert wb[wb.sheetnames[0]]["C9"].value == "T33U-26040025-0002"


def test_booking_body_validation_incomplete_po_uses_all_delivery_list_candidates() -> None:
    pn = "1010300202002T01"
    content = _workbook_bytes([_sil_fuca_row(po="T33U-26040025", pn=pn, quantity=20000)])
    client = _FakeSilFucaDeliveryClient(
        all_records=(
            _delivery_record(po="T33U-26040025-0002", pn=pn, delivery_quantity=20000),
        ),
    )

    preview = build_body_validation_preview(
        content,
        filename="sil.xlsx",
        apply_fixes=True,
        enable_dynamic_checks=True,
        sil_fuca_delivery_client=client,
        query_date=date(2026, 6, 29),
    )

    row = preview.rows[0]
    assert client.queries == []
    assert client.all_call_count == 1
    assert row.values["PO_No"] == "T33U-26040025-0002"
    assert row.delivery_match_status == "ok"
    assert row.correction_kind_for("PO_No") == "sil_fuca_delivery_po"
    assert "PO_No" in row.fixed_fields


def test_booking_body_validation_sil_fuca_delivery_flags_missing_record() -> None:
    pn = "1010300202002T01"
    content = _workbook_bytes([_sil_fuca_row(po="K33U-24090083-0001", pn=pn, quantity=20000)])
    client = _FakeSilFucaDeliveryClient(
        new_errors={
            ("K33U-24090083-0001", pn): (
                "K33U-24090083-0001 匹配不上周期交货清单，请检查是否有上传",
            )
        }
    )

    preview = build_body_validation_preview(
        content,
        filename="sil.xlsx",
        apply_fixes=True,
        enable_dynamic_checks=True,
        sil_fuca_delivery_client=client,
        query_date=date(2026, 6, 29),
    )

    assert preview.issue_count == 1
    assert preview.issues[0].correction_kind == "sil_fuca_delivery_missing"
    assert "匹配不上周期交货清单" in preview.issues[0].message
    assert "ASN" in preview.rows[0].issue_fields
    assert "ASN" in preview.rows[0].source_issue_fields
    assert preview.rows[0].delivery_match_status == "error"


def test_booking_body_validation_sil_fuca_delivery_checks_quantity_and_date() -> None:
    pn = "1010300202002T01"
    content = _workbook_bytes(
        [
            _sil_fuca_row(po="T33U-26040025-0002", pn=pn, quantity=21000),
            _sil_fuca_row(po="K33U-26040025-0002", pn=pn, quantity=1000),
        ]
    )
    client = _FakeSilFucaDeliveryClient(
        new_records={
            ("T33U-26040025-0002", pn): (
                _delivery_record(po="T33U-26040025-0002", pn=pn, delivery_quantity=20000),
            ),
            ("K33U-26040025-0002", pn): (
                _delivery_record(po="K33U-26040025-0002", pn=pn, delivery_quantity=20000, delivery_date="2026-06-29"),
            ),
        }
    )

    preview = build_body_validation_preview(
        content,
        filename="sil.xlsx",
        apply_fixes=True,
        enable_dynamic_checks=True,
        sil_fuca_delivery_client=client,
        query_date=date(2026, 6, 29),
    )

    messages = [issue.message for issue in preview.issues if issue.field_code == "ASN"]
    assert any("数量不足" in message and "周期数量 20000" in message for message in messages)
    assert any("交货日期 2026-06-29 不晚于当前查询日期" in message for message in messages)


def test_booking_body_validation_route_marks_error_cells() -> None:
    content = _workbook_bytes(
        [
            [
                1,
                "",
                "BADPO",
                "PN-1",
                "IC",
                0,
                "PCS",
                0,
                0,
                0,
                0,
                "",
                "INV #1",
                "bad-date",
                "",
                "MARS",
                "",
                "",
                "",
                "BAD",
                "",
                "NA",
                "",
                10,
                3,
                "",
                "",
            ]
        ]
    )
    client = TestClient(booking_app)

    response = client.post(
        "/modules/booking/body-validation",
        data={"apply_fixes": "0"},
        files={
            "booking_file": (
                "supplier_booking.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert "源数据导入表" in response.text
    assert "screening-cell-error" in response.text
    assert "PO No. 格式应类似" in response.text


def test_booking_body_validation_export_downloads_corrected_workbook(monkeypatch, tmp_path) -> None:
    monkeypatch.setattr(booking_routes, "BODY_VALIDATION_EXPORT_DIR", tmp_path)
    monkeypatch.setattr(booking_routes, "BODY_VALIDATION_UPLOAD_DIR", tmp_path / "uploads")
    content = _workbook_bytes(
        [
            _valid_row(made_date="2026-01-12 AND 2026-01-26", per_box="2K+2K"),
            _valid_row(made_date=datetime(2026, 3, 23), per_box=5000),
        ]
    )
    client = TestClient(booking_app)

    response = client.post(
        "/modules/booking/body-validation",
        data={"apply_fixes": "0"},
        files={
            "booking_file": (
                "supplier booking.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert "/modules/booking/body-validation/export/" not in response.text
    session_match = re.search(r'name="validation_session_id" value="([0-9a-f]{32})"', response.text)
    assert session_match

    suggestion_response = client.post(
        "/modules/booking/body-validation",
        data={"apply_fixes": "1", "validation_session_id": session_match.group(1)},
    )

    assert suggestion_response.status_code == 200
    assert "/modules/booking/body-validation/export/" not in suggestion_response.text

    export_response = client.post(
        "/modules/booking/body-validation",
        data={
            "apply_fixes": "1",
            "validation_session_id": session_match.group(1),
            "confirm_export": "1",
            "manual_values_json": json.dumps(
                {
                    "9": {"madeDate": "2026-01-12", "per_box": "4000"},
                    "10": {"madeDate": "2026-03-23"},
                }
            ),
        },
    )

    assert export_response.status_code == 200
    match = re.search(r'href="([^"]*/modules/booking/body-validation/export/[^"]+)"', export_response.text)
    assert match

    download = client.get(match.group(1))
    assert download.status_code == 200
    assert download.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    wb = load_workbook(BytesIO(download.content))
    ws = wb[wb.sheetnames[0]]
    assert ws["N9"].value == datetime(2026, 1, 12)
    assert ws["N9"].number_format == "yyyy-mm-dd"
    assert ws["N10"].value == datetime(2026, 3, 23)
    assert ws["N10"].number_format == "yyyy-mm-dd"
    assert str(ws["Y9"].value) == "4000"


def test_booking_body_validation_can_start_from_generated_booking_preview(monkeypatch, tmp_path) -> None:
    monkeypatch.setattr(booking_routes, "BODY_VALIDATION_UPLOAD_DIR", tmp_path / "uploads")
    output_path = tmp_path / "generated_booking.xlsx"
    output_path.write_bytes(_workbook_bytes([_valid_row()]))
    preview = BookingPreview(
        session_id="generated-preview",
        supplier="SIL-FUCA",
        source_filename="source.eml",
        pack_filename="pack.xlsx",
        rows=[{"PO No. *": "C33C-25040701-0001"}],
        columns=["PO No. *"],
    )
    booking_routes.SESSION_STORE["generated-preview"] = {"booking_preview": preview}
    monkeypatch.setattr(booking_routes, "write_booking_output", lambda _preview: output_path)
    client = TestClient(booking_app)

    response = client.get("/modules/booking/body-validation/from-preview/generated-preview")

    assert response.status_code == 200
    assert "源数据导入表" in response.text
    assert "generated_booking.xlsx" in response.text
    assert re.search(r'name="validation_session_id" value="[0-9a-f]{32}"', response.text)


def test_booking_body_validation_extension_upload_redirects_to_session(monkeypatch, tmp_path) -> None:
    monkeypatch.setattr(booking_routes, "BODY_VALIDATION_UPLOAD_DIR", tmp_path / "uploads")
    content = _workbook_bytes([_valid_row()])
    client = TestClient(booking_app, follow_redirects=False)

    response = client.post(
        "/modules/booking/body-validation/extension-upload",
        files={
            "booking_file": (
                "supplier_booking.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 303
    assert re.fullmatch(r"/modules/booking/body-validation/session/[0-9a-f]{32}", response.headers["location"])

    result_response = TestClient(booking_app).get(response.headers["location"])
    assert result_response.status_code == 200
    assert "supplier_booking.xlsx" in result_response.text
    assert "源数据导入表" in result_response.text
