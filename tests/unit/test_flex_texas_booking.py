from __future__ import annotations

import re
from email.message import EmailMessage
from email import policy
from email.parser import BytesParser
from pathlib import Path
from types import SimpleNamespace

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from PIL import Image

from app.modules.booking.schemas import BookingPreview
from booking_web_app import app as booking_app
from app.modules.booking import routes as booking_routes
from app.modules.booking.flex_texas import (
    TexasAwbInfo,
    TexasBookingLine,
    TexasEmailInfo,
    build_texas_validation_sections,
    calculate_texas_cbm,
    classify_texas_pdf_pages,
    export_flex_texas_source_pdf_tiff,
    is_system_export_pdf,
    join_invoice_and_content_list_by_waybill,
    parse_flex_texas_eml,
    parse_texas_awb_layout,
    parse_texas_commercial_invoice_layout,
    parse_texas_content_list_layout,
    parse_texas_email_body,
)
from app.modules.booking.mail_builder import generate_flex_texas_booking_reply_eml
from app.modules.booking.rules.registry import get_supplier_names
from booking_store import build_booking_preview, write_booking_workbook


SAMPLE_DIR = Path.home() / "Desktop" / "新建文件夹"


EXPECTED = {
    "4711941716": {
        "hblNo": "4711941716",
        "mblNo": "16007966630",
        "deliveryToHubDate": "2026/06/11",
        "lines": [
            {
                "invoiceNo": "5536326997",
                "invoiceDate": "2026/06/08",
                "tiWaybillNo": "431826703",
                "poNo": "NPI003983",
                "customerPartNo": "FPS-213-D0001022-103",
                "madeIn": "US",
                "cartons": 1,
                "grossWeight": 1.7,
                "boxSize": "400X400X120MM",
                "cbm": 0.01,
                "quantity": 474,
                "price": 1.95,
                "totalAmount": 924.3,
            }
        ],
    },
    "4397543788": {
        "hblNo": "4397543788",
        "mblNo": "11206432064",
        "deliveryToHubDate": "2026/06/11",
        "lines": [
            {
                "invoiceNo": "5536272465",
                "invoiceDate": "2026/06/06",
                "tiWaybillNo": "431808752",
                "poNo": "NPI003826",
                "customerPartNo": "FPS-213-D0000834-101",
                "madeIn": "US",
                "cartons": 1,
                "grossWeight": 1.43,
                "boxSize": "400X400X120MM",
                "cbm": 0.01,
                "quantity": 2500,
                "price": 0.099,
                "totalAmount": 247.5,
            }
        ],
    },
    "4397545049": {
        "hblNo": "4397545049",
        "mblNo": "11206432064",
        "deliveryToHubDate": "2026/06/11",
        "lines": [
            {
                "invoiceNo": "5536341426",
                "invoiceDate": "2026/06/08",
                "tiWaybillNo": "431822191",
                "poNo": "J50018294",
                "customerPartNo": "FPS-213-D0000259-101",
                "madeIn": "JP",
                "cartons": 1,
                "grossWeight": 0.91,
                "boxSize": "240X240X100MM",
                "cbm": 0.01,
                "quantity": 6000,
                "price": 0.13,
                "totalAmount": 780,
            }
        ],
    },
    "4711940588": {
        "hblNo": "4711940588",
        "mblNo": "16007966232",
        "deliveryToHubDate": "2026/06/07",
        "lines": [
            {
                "invoiceNo": "5536104817",
                "invoiceDate": "2026/06/04",
                "tiWaybillNo": "431737534",
                "poNo": "J50011399",
                "customerPartNo": "SNBH-GEP0202038",
                "madeIn": "US",
                "cartons": 1,
                "grossWeight": 1.63,
                "boxSize": "300X240X240MM",
                "cbm": 0.01,
                "quantity": 18000,
                "price": 0.039,
                "totalAmount": 702,
            },
            {
                "invoiceNo": "5536104818",
                "invoiceDate": "2026/06/04",
                "tiWaybillNo": "431762478",
                "poNo": "J50012492",
                "customerPartNo": "GECH-720051759-888-0A",
                "madeIn": "DE",
                "cartons": 1,
                "grossWeight": 0.5,
                "boxSize": "240X240X100MM",
                "cbm": 0.01,
                "quantity": 3000,
                "price": 0.281,
                "totalAmount": 843,
            },
        ],
        "totals": {
            "cartons": 2,
            "pallet": 0,
            "grossWeight": 2.13,
            "cbm": 0.02,
            "quantity": 21000,
            "totalAmount": 1545,
        },
    },
}


def _sample_eml(hbl_no: str) -> Path:
    matches = list(SAMPLE_DIR.glob(f"*HB#{hbl_no}.eml"))
    if not matches:
        pytest.skip(f"Flex-Texas sample EML {hbl_no} is not available on this machine.")
    return matches[0]


def _sample_pdf_payload(hbl_no: str) -> bytes:
    message = BytesParser(policy=policy.default).parsebytes(_sample_eml(hbl_no).read_bytes())
    for part in message.walk():
        filename = part.get_filename() or ""
        if filename.lower().endswith(".pdf"):
            payload = part.get_payload(decode=True)
            if payload:
                return payload
    pytest.fail(f"Sample EML {hbl_no} did not contain a PDF attachment.")


def test_flex_texas_supplier_is_registered() -> None:
    assert "FLEX-TEXAS" in get_supplier_names()


def test_booking_page_exposes_flex_texas_eml_mode() -> None:
    client = TestClient(booking_app)

    response = client.get("/modules/booking?supplier=FLEX-TEXAS")

    assert response.status_code == 200
    assert 'value="FLEX-TEXAS"' in response.text
    assert 'value="SIL-FUCA"' not in response.text
    assert 'data-eml-pdf="true"' in response.text
    assert 'name="auto_generate"' in response.text
    assert "选择 Flex-Texas 原始邮件 .eml" in response.text
    assert "TIF 核对" in response.text


def test_general_booking_page_still_exposes_sil_supplier() -> None:
    client = TestClient(booking_app)

    response = client.get("/modules/booking")

    assert response.status_code == 200
    assert 'value="SIL-FUCA"' in response.text


def test_booking_lock_supplier_env_hides_sil(monkeypatch) -> None:
    monkeypatch.setenv("BOOKING_LOCK_SUPPLIER", "FLEX-TEXAS")
    client = TestClient(booking_app)

    response = client.get("/modules/booking")

    assert response.status_code == 200
    assert 'value="FLEX-TEXAS"' in response.text
    assert 'value="SIL-FUCA"' not in response.text


def test_booking_preview_opens_flex_texas_pdf_as_tiff_instead_of_thumbnail(monkeypatch) -> None:
    eml_path = _sample_eml("4711941716")
    parsed = parse_flex_texas_eml(eml_path)
    client = TestClient(booking_app)
    opened_paths: list[Path] = []

    def fake_open_review_tiff(path: Path) -> None:
        opened_paths.append(path)

    monkeypatch.setattr(booking_routes, "open_review_tiff", fake_open_review_tiff)

    response = client.post(
        "/modules/booking/preview",
        data={"supplier": "FLEX-TEXAS", "auto_generate": "1"},
        files={"customer_eml": (eml_path.name, eml_path.read_bytes(), "message/rfc822")},
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/html")
    assert not response.headers.get("content-disposition")
    assert "即将填入 booking form 的数值" in response.text
    assert "字段校验汇总" in response.text
    assert "MAWB / MBLNO" in response.text
    assert "HAWB / H.B/LNO" in response.text
    assert '<details class="booking-validation-section is-ok">' in response.text
    assert '<details class="booking-validation-item is-ok">' in response.text
    assert 'action="/modules/booking/flex-texas-reply-mail/' in response.text
    assert 'name="tms_pdf_file"' in response.text
    assert 'name="body_text"' in response.text
    assert 'class="booking-field-records"' in response.text
    assert 'class="booking-table"' not in response.text
    assert "data-auto-tiff-download" in response.text
    assert "Windows 照片查看器" in response.text
    assert 'data-pdf-review' not in response.text
    assert "/modules/booking/pdf-preview/" not in response.text
    assert "NPI003983" in response.text
    assert "5536326997" in response.text
    assert len(opened_paths) == 1
    assert opened_paths[0].suffix.lower() == ".tif"
    assert opened_paths[0].is_file()
    with Image.open(opened_paths[0]) as image:
        assert image.format == "TIFF"
        assert getattr(image, "n_frames", 1) == parsed.pdf_page_count


    generate_match = re.search(r'href="/modules/booking/generate/([^"]+)"', response.text)
    assert generate_match
    generate_response = client.get(f"/modules/booking/generate/{generate_match.group(1)}")
    assert generate_response.status_code == 200
    assert generate_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment;" in generate_response.headers["content-disposition"]
    assert generate_response.content.startswith(b"PK")

    post_generate_response = client.post(f"/modules/booking/generate/{generate_match.group(1)}")
    assert post_generate_response.status_code == 200
    assert post_generate_response.content.startswith(b"PK")


def test_booking_preview_exposes_tiff_download_when_server_is_not_windows(monkeypatch) -> None:
    eml_path = _sample_eml("4711941716")
    client = TestClient(booking_app)
    opened_paths: list[Path] = []

    monkeypatch.setattr(booking_routes.os, "name", "posix")
    monkeypatch.setattr(booking_routes, "open_review_tiff", lambda path: opened_paths.append(path))

    response = client.post(
        "/modules/booking/preview",
        data={"supplier": "FLEX-TEXAS", "auto_generate": "1"},
        files={"customer_eml": (eml_path.name, eml_path.read_bytes(), "message/rfc822")},
    )

    assert response.status_code == 200
    assert opened_paths == []
    tiff_match = re.search(r'href="(/modules/booking/flex-texas-review-tiff/[^"]+)"', response.text)
    assert tiff_match
    assert "data-auto-tiff-download" in response.text

    tiff_response = client.get(tiff_match.group(1))

    assert tiff_response.status_code == 200
    assert tiff_response.headers["content-type"].startswith("image/tiff")
    assert "attachment;" in tiff_response.headers["content-disposition"]
    assert tiff_response.content[:2] in {b"II", b"MM"}


def test_flex_texas_review_tiff_download_route_returns_attachment(tmp_path: Path) -> None:
    client = TestClient(booking_app)
    session_id = "direct-tiff-route"
    tiff_path = tmp_path / "review.tif"
    Image.new("RGB", (20, 20), "white").save(tiff_path, format="TIFF")
    booking_routes.SESSION_STORE[session_id] = {"booking_pdf_tiff_path": str(tiff_path)}

    response = client.get(f"/modules/booking/flex-texas-review-tiff/{session_id}")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("image/tiff")
    assert "attachment;" in response.headers["content-disposition"]
    assert response.content[:2] in {b"II", b"MM"}


def test_booking_xlsx_download_route_returns_attachment(monkeypatch, tmp_path: Path) -> None:
    client = TestClient(booking_app)
    session_id = "direct-xlsx-route"
    output_path = tmp_path / "booking.xlsx"
    output_path.write_bytes(b"PK\x03\x04synthetic-xlsx")
    preview = BookingPreview(
        session_id=session_id,
        supplier="FLEX-TEXAS",
        source_filename="source.eml",
        pack_filename="source.pdf",
        rows=[{"P/N": "TEST"}],
        columns=["P/N"],
    )
    booking_routes.SESSION_STORE[session_id] = {"booking_preview": preview}
    monkeypatch.setattr(booking_routes, "write_booking_output", lambda _preview: output_path)

    response = client.get(f"/modules/booking/generate/{session_id}")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment;" in response.headers["content-disposition"]
    assert response.content.startswith(b"PK")


def test_export_flex_texas_pdf_tiff_ignores_system_export_pdf(tmp_path: Path) -> None:
    output_path = export_flex_texas_source_pdf_tiff(_sample_eml("4711941716"), tmp_path / "review.tif")

    assert output_path.is_file()
    with Image.open(output_path) as image:
        assert image.format == "TIFF"
        assert getattr(image, "n_frames", 1) == parse_flex_texas_eml(_sample_eml("4711941716")).pdf_page_count


def test_generate_flex_texas_reply_eml_threads_to_original(tmp_path: Path) -> None:
    original = EmailMessage(policy=policy.SMTP)
    original["Subject"] = "Flextronics c/o Smooth - Notification of New Shipment Arrival - HB#4397543788"
    original["From"] = "Customer <customer@example.com>"
    original["Reply-To"] = "booking-desk@example.com"
    original["Message-ID"] = "<source-message@example.com>"
    original["References"] = "<thread-root@example.com>"
    original.set_content("source mail")
    customer_eml_path = tmp_path / "source.eml"
    customer_eml_path.write_bytes(original.as_bytes())
    tms_pdf_path = tmp_path / "TMS-4397543788.pdf"
    tms_pdf_path.write_bytes(b"%PDF-1.4\n%test\n")
    output_path = tmp_path / "reply.eml"

    subject = generate_flex_texas_booking_reply_eml(
        preview=SimpleNamespace(supplier="FLEX-TEXAS", mawb_no="11206432064", hbl_no="4397543788"),
        customer_eml_path=customer_eml_path,
        tms_pdf_path=tms_pdf_path,
        output_path=output_path,
    )

    generated = BytesParser(policy=policy.default).parsebytes(output_path.read_bytes())
    assert subject == "Re: Flextronics c/o Smooth - Notification of New Shipment Arrival - HB#4397543788"
    assert generated["X-Unsent"] == "1"
    assert generated["To"] == "booking-desk@example.com"
    assert generated["In-Reply-To"] == "<source-message@example.com>"
    assert generated["References"] == "<thread-root@example.com> <source-message@example.com>"
    assert "completed TMS warehouse-entry PDF" in generated.get_body(preferencelist=("plain",)).get_content()
    attachments = [part for part in generated.walk() if part.get_content_disposition() == "attachment"]
    assert [part.get_filename() for part in attachments] == ["TMS-4397543788.pdf"]


def test_parse_texas_email_body_extracts_table_values() -> None:
    text = """
HAWB
File
MAWB
Client
PKG
Weight
ETA
4711941716
2912444453
160-07966630
Flextronics
001 PCS
2.0 K
CX658/11 Jun
"""

    parsed = parse_texas_email_body(
        text,
        email_date="Thu, 11 Jun 2026 01:52:42 +0000",
        subject="Flextronics c/o Smooth - Notification of New Shipment Arrival - HB#4711941716",
    )

    assert parsed.hbl_no == "4711941716"
    assert parsed.subject_hbl_no == "4711941716"
    assert parsed.mbl_no == "16007966630"
    assert parsed.cartons == 1
    assert parsed.gross_weight == 2.0
    assert parsed.delivery_to_hub_date == "2026/06/11"


def test_classify_texas_pdf_pages_and_ignore_comments_page() -> None:
    pages = classify_texas_pdf_pages(_sample_pdf_payload("4711940588"), source_file="4711940588.pdf")

    assert [page.page_type for page in pages] == [
        "air_waybill",
        "commercial_invoice_item",
        "commercial_invoice_comments",
        "commercial_invoice_item",
        "commercial_invoice_comments",
        "content_list",
        "content_list",
    ]
    assert sum(page.page_type == "commercial_invoice_item" for page in pages) == 2


def test_parse_texas_awb_layout_extracts_header_and_dimensions() -> None:
    pages = classify_texas_pdf_pages(_sample_pdf_payload("4711940588"), source_file="4711940588.pdf")
    awb = parse_texas_awb_layout(pages[0])

    assert awb.hbl_no == "4711940588"
    assert awb.mbl_no == "16007966232"
    assert awb.carrier_code == "EI"
    assert awb.dimensions == "1CTN@30X24X24;1CTN@25X24X11"


def test_parse_invoice_and_content_list_layouts() -> None:
    pages = classify_texas_pdf_pages(_sample_pdf_payload("4711941716"), source_file="4711941716.pdf")
    invoice = parse_texas_commercial_invoice_layout(pages[1])[0]
    content = parse_texas_content_list_layout(pages[3])[0]

    assert invoice.invoice_no == "5536326997"
    assert invoice.invoice_date == "2026/06/08"
    assert invoice.ti_waybill_no == "431826703"
    assert invoice.customer_part_no == "FPS-213-D0001022-103"
    assert invoice.made_in == "US"
    assert invoice.quantity == 474
    assert invoice.price == 1.95
    assert invoice.total_amount == 924.3
    assert content.waybill_no == "431826703"
    assert content.po_no == "NPI003983"
    assert content.customer_part_no == "FPS-213-D0001022-103"
    assert content.box_size == "400X400X120MM"
    assert content.gross_weight == 1.7
    assert content.cbm == 0.01


def test_join_invoice_and_content_list_by_waybill_not_page_order() -> None:
    pages = classify_texas_pdf_pages(_sample_pdf_payload("4711940588"), source_file="4711940588.pdf")
    invoices = [
        invoice
        for page in pages
        if page.page_type == "commercial_invoice_item"
        for invoice in parse_texas_commercial_invoice_layout(page)
    ]
    content_rows = [
        row
        for page in pages
        if page.page_type == "content_list"
        for row in parse_texas_content_list_layout(page)
    ]

    lines, warnings = join_invoice_and_content_list_by_waybill(invoices, content_rows)

    assert warnings == []
    assert lines[0].invoice_no == "5536104817"
    assert lines[0].ti_waybill_no == "431737534"
    assert lines[0].box_size == "300X240X240MM"
    assert lines[0].po_no == "J50011399"
    assert lines[1].invoice_no == "5536104818"
    assert lines[1].ti_waybill_no == "431762478"
    assert lines[1].box_size == "240X240X100MM"
    assert lines[1].po_no == "J50012492"


@pytest.mark.parametrize("hbl_no", sorted(EXPECTED))
def test_parse_flex_texas_sample_expected_extraction(hbl_no: str) -> None:
    expected = EXPECTED[hbl_no]

    parsed = parse_flex_texas_eml(_sample_eml(hbl_no))

    assert parsed.hbl_no == expected["hblNo"]
    assert parsed.mbl_no == expected["mblNo"]
    assert parsed.delivery_to_hub_date == expected["deliveryToHubDate"]
    assert parsed.transport_company == "EI"
    assert parsed.delivery_party == "VIA AIR"
    assert len(parsed.lines) == len(expected["lines"])
    for line, expected_line in zip(parsed.lines, expected["lines"], strict=True):
        assert line.invoice_no == expected_line["invoiceNo"]
        assert line.invoice_date == expected_line["invoiceDate"]
        assert line.ti_waybill_no == expected_line["tiWaybillNo"]
        assert line.po_no == expected_line["poNo"]
        assert line.customer_part_no == expected_line["customerPartNo"]
        assert line.made_in == expected_line["madeIn"]
        assert line.cartons == expected_line["cartons"]
        assert line.gross_weight == pytest.approx(expected_line["grossWeight"])
        assert line.box_size == expected_line["boxSize"]
        assert line.cbm == pytest.approx(expected_line["cbm"])
        assert line.quantity == expected_line["quantity"]
        assert line.price == pytest.approx(expected_line["price"])
        assert line.currency == "USD"
        assert line.total_amount == pytest.approx(expected_line["totalAmount"])
        assert line.part_description == "MATERIAL"
        assert line.tray_type == "0"
    if "totals" in expected:
        assert parsed.totals == expected["totals"]


def test_fill_booking_template_texas(tmp_path: Path) -> None:
    if not (SAMPLE_DIR / "smooth booking template.xlsx").exists():
        pytest.skip("smooth booking template.xlsx is not available on this machine.")
    preview = build_booking_preview(
        session_id="flex-texas-test",
        supplier="FLEX-TEXAS",
        source_path=_sample_eml("4711940588"),
    )

    output_path = write_booking_workbook(preview, tmp_path)
    ws = load_workbook(output_path, data_only=False).active

    assert ws["K4"].value == "16007966232"
    assert ws["N4"].value == "4711940588"
    assert ws["G4"].value == "2026/06/07"
    assert ws["K5"].value == "VIA AIR"
    assert ws["O5"].value == "EI"
    assert ws["B10"].value == "J50011399"
    assert ws["D10"].value == "5536104817"
    assert ws["F10"].value == "SNBH-GEP0202038"
    assert ws["B11"].value == "J50012492"
    assert ws["D11"].value == "5536104818"
    assert ws["F11"].value == "GECH-720051759-888-0A"
    assert ws["I20"].value == 2
    assert ws["J20"].value == 0
    assert ws["K20"].value == pytest.approx(2.13)
    assert ws["L20"].value == pytest.approx(0.02)
    assert ws["M20"].value == 21000
    assert ws["P20"].value == 1545
    assert "A20:F20" in {str(item) for item in ws.merged_cells.ranges}


def test_do_not_parse_system_export_pdf_as_input() -> None:
    assert is_system_export_pdf("RH2604405 260611.pdf")


def test_texas_cbm_uses_existing_truncate_style() -> None:
    assert calculate_texas_cbm("400X400X120MM") == 0.01
    assert calculate_texas_cbm("240X240X100MM") == 0.01
    assert calculate_texas_cbm("300X240X240MM") == 0.01


def test_flex_texas_validation_sections_pass_for_real_sample() -> None:
    parsed = parse_flex_texas_eml(_sample_eml("4397543788"))

    items = [item for section in parsed.validation_sections for item in section.items]

    assert items
    assert {item.status for item in items} == {"ok"}


def test_flex_texas_validation_sections_warn_on_confirmed_rules() -> None:
    sections = build_texas_validation_sections(
        final_mbl_no="11206432065",
        final_hbl_no="123456789",
        awb=TexasAwbInfo(hbl_no="123456789", mbl_no="112-06432065"),
        email_info=TexasEmailInfo(hbl_no="123456789", mbl_no="11206432065", subject_hbl_no="123456789"),
        lines=[
            TexasBookingLine(
                invoice_no="553627246",
                invoice_date="2026/06/06",
                ti_waybill_no="431808752",
                po_no="BAD003826",
                invoice_customer_po_raw="BAD003826",
                customer_part_no="BAD-213-D0000834-101",
                made_in="US",
                cartons=1,
                pallet=0,
                gross_weight=1.43,
                box_size="400X400X120MM",
                cbm=0.01,
                quantity=3,
                price=2,
                currency="USD",
                total_amount=7,
            )
        ],
    )
    items = {(item.field, item.row_no): item for section in sections for item in section.items}
    identity_items = {item.field: item for section in sections for item in section.items if item.row_no is None}

    assert identity_items["MAWB / MBLNO"].status == "error"
    assert identity_items["HAWB / H.B/LNO"].status == "error"
    assert items[("Invoice No.", 1)].status == "warning"
    assert items[("PO No.", 1)].status == "warning"
    assert items[("Customer Part No.", 1)].status == "warning"
    assert items[("Total Amount", 1)].status == "warning"
