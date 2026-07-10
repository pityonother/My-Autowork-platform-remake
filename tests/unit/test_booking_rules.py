from __future__ import annotations

from datetime import date
from email import policy
from email.message import EmailMessage
from email.parser import BytesParser
from functools import partial
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

import pytest
from fastapi import UploadFile
from openpyxl import Workbook, load_workbook

import booking_store
from app.modules.booking import mail_builder, service as booking_service
from app.modules.booking.legacy_adapter import BookingPreview, build_booking_preview, write_booking_workbook
from app.modules.booking.mail_builder import extract_sil_warehouse_no, replace_mail_template_values
from app.modules.booking.rules.registry import SUPPLIER_RULES, get_supplier_names
from app.shared.uploads import UploadValidationError, save_upload
from booking_rules import vc_dzyq


def test_booking_rule_registry_exposes_expected_suppliers() -> None:
    assert "SIL-FUCA" in get_supplier_names()
    assert "SIL-WEIKENG" in get_supplier_names()
    assert "VC_DZYQ" in get_supplier_names()


def test_sil_fuca_keeps_customer_po_without_default_item_suffix(tmp_path) -> None:
    source_path = tmp_path / "CCIXLS_sil_fuca.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "detail"
    ws.append(
        [
            "Customer PO",
            "Customer Part Number",
            "HS Desc",
            "Quantity",
            "Net Weight",
            "Gross Weight",
            "CofO",
            "MFR Name",
        ]
    )
    ws.append(["T33U-26040113", "1010135800000", "IC", 20000, 1.2, 1.4, "CHINA", "TI"])
    wb.save(source_path)

    preview = build_booking_preview(session_id="sil-no-default-item", supplier="SIL-FUCA", source_path=source_path)

    assert preview.can_generate
    assert preview.rows[0]["订单号"] == "T33U-26040113"


def test_supplier_specific_template_does_not_fallback_to_default(tmp_path) -> None:
    rule = SimpleNamespace(
        TEMPLATE_CANDIDATES=[tmp_path / "smooth booking template.xlsx"],
        REQUIRE_TEMPLATE_CANDIDATE=True,
        FLEX_TEXAS_TEMPLATE_NAME="smooth booking template.xlsx",
    )

    try:
        booking_store.get_default_booking_template(rule)
    except FileNotFoundError as exc:
        message = str(exc)
    else:
        raise AssertionError("supplier-specific templates must not fall back to booking_template_zh.xlsx")

    assert "smooth booking template.xlsx" in message
    assert str(tmp_path / "smooth booking template.xlsx") in message


def test_weikeng_total_box_count_is_written_only_on_first_row() -> None:
    rule = SUPPLIER_RULES["SIL-WEIKENG"]

    extras, warnings = rule.post_process(
        [{"_summary_box_count": 12}, {"_summary_box_count": 12}],
        [],
    )

    assert warnings == []
    assert extras == [{"Total Box Count": 12}, {"Total Box Count": 0}]


def test_weikeng_pallet_count_is_zero_after_first_row(tmp_path) -> None:
    source_path = tmp_path / "weikeng_packing.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "packing"
    ws.append(
        [
            "採購商訂單號",
            "訂單項次",
            "產地",
            "品牌",
            "供應商發票號",
            "箱號",
            "數量",
            "數量單位",
            "毛重",
            "淨重",
            "材積",
            "DATECODE",
            "採購商物料號",
            "LOTNO",
            "IC屬性",
            "最小包装数量",
            "每箱数量",
            "總箱數",
        ]
    )
    ws.append(["C33C-26010025", "1", "US", "TI", "INV-1", "1", 100, "PCS", 1.2, 1.0, "10X10X10", "2601", "PART-001", "LOT1", "IC", 50, 100, ""])
    ws.append(["C33C-26010025", "2", "US", "TI", "INV-1", "2", 200, "PCS", 2.2, 2.0, "10X10X10", "2601", "PART-002", "LOT2", "IC", 50, 100, ""])
    ws.append(["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", 2])
    wb.save(source_path)

    preview = build_booking_preview(session_id="weikeng-pallet-test", supplier="SIL-WEIKENG", source_path=source_path)

    assert preview.can_generate
    assert "板数" in preview.columns
    assert preview.rows[0]["板数"] == ""
    assert preview.rows[1]["板数"] == 0


def test_sil_warehouse_mail_helpers_replace_mawb_and_warehouse_no(tmp_path) -> None:
    warehouse_file = tmp_path / "SIL26040490 入仓纸.pdf"

    assert extract_sil_warehouse_no(warehouse_file) == "SIL26040490"
    assert (
        replace_mail_template_values("MAWB# 12345678901<br>SIL26040000", "98765432109", "SIL26040490")
        == "MAWB# 98765432109<br>SIL26040490"
    )


def test_sil_warehouse_mail_removes_obsolete_holiday_notice() -> None:
    old_html = (
        '<p class="MsoNormal"><span>香港仓：5月1日&amp;5月5日（全日休息）</span></p>'
        "<p>香港仓库上下班时间：上午09:00-12:00</p>"
    )
    old_plain = "香港仓：5月1日&5月5日（全日休息）\r\n香港仓库上下班时间：上午09:00-12:00\r\n"

    cleaned_html = mail_builder.remove_obsolete_sil_holiday_notice(old_html)
    cleaned_plain = mail_builder.remove_obsolete_sil_holiday_notice(old_plain)

    assert "香港仓：5月1日" not in cleaned_html
    assert "香港仓：5月1日" not in cleaned_plain
    assert "香港仓库上下班时间" in cleaned_html
    assert "香港仓库上下班时间" in cleaned_plain


def test_sil_warehouse_template_falls_back_to_packaged_default(monkeypatch, tmp_path) -> None:
    monkeypatch.setattr(
        mail_builder,
        "SIL_FUCA_WAREHOUSE_TEMPLATE_JSON",
        tmp_path / "missing-runtime-template" / "template.json",
    )

    template = mail_builder.load_sil_fuca_warehouse_template()
    base_dir = Path(str(template["_template_base_dir"]))

    assert template["html"]
    assert template["assets"]
    assert mail_builder.resolve_template_asset_path(template["assets"][0], base_dir) is not None


def test_generate_sil_warehouse_eml_uses_html_default_template(monkeypatch, tmp_path) -> None:
    monkeypatch.setattr(mail_builder, "RUNTIME_DIR", tmp_path / "runtime")
    monkeypatch.setattr(
        mail_builder,
        "SIL_FUCA_WAREHOUSE_TEMPLATE_JSON",
        tmp_path / "missing-runtime-template" / "template.json",
    )

    customer_message = EmailMessage()
    customer_message["Subject"] = "MAWB# 98765432109"
    customer_message.set_content("customer body")
    customer_eml = tmp_path / "customer.eml"
    customer_eml.write_bytes(customer_message.as_bytes(policy=policy.default))

    warehouse_file = tmp_path / "SIL26040490.pdf"
    warehouse_file.write_bytes(b"warehouse")
    preview = SimpleNamespace(
        supplier="SIL-FUCA",
        mawb_no="",
        email_subject="MAWB# 98765432109",
        session_id="rich-template-test",
    )
    output_path = tmp_path / "warehouse.eml"

    mail_builder.generate_sil_fuca_warehouse_eml(
        preview=preview,
        customer_eml_path=customer_eml,
        warehouse_file_path=warehouse_file,
        output_path=output_path,
    )

    message = BytesParser(policy=policy.default).parsebytes(output_path.read_bytes())
    html_part = message.get_body(preferencelist=("html",))
    plain_part = message.get_body(preferencelist=("plain",))

    assert html_part is not None
    assert plain_part is not None
    assert "98765432109" in html_part.get_content()
    assert "SIL26040490" in html_part.get_content()
    assert "香港仓：5月1日" not in html_part.get_content()
    assert "香港仓：5月1日" not in plain_part.get_content()


@pytest.mark.parametrize(
    ("warehouse_filename", "payload", "expected_content_type"),
    [
        ("SIL26040490.pdf", b"%PDF-1.4\n%warehouse\n", "application/pdf"),
        ("SIL26040490.xls", b"legacy warehouse workbook", "application/vnd.ms-excel"),
        (
            "SIL26040490.xlsx",
            b"PK\x03\x04warehouse workbook",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
    ],
)
def test_write_sil_warehouse_mail_accepts_pdf_and_spreadsheets_as_attachments(
    monkeypatch,
    tmp_path: Path,
    warehouse_filename: str,
    payload: bytes,
    expected_content_type: str,
) -> None:
    monkeypatch.setattr(
        booking_service,
        "save_upload",
        partial(save_upload, upload_root=tmp_path / "uploads"),
    )
    monkeypatch.setattr(booking_service, "OUTPUT_DIR", tmp_path / "outputs")
    monkeypatch.setattr(mail_builder, "RUNTIME_DIR", tmp_path / "runtime")
    monkeypatch.setattr(
        mail_builder,
        "load_sil_fuca_warehouse_template",
        lambda: {"html": "", "plain": "", "assets": []},
    )

    customer_message = EmailMessage(policy=policy.SMTP)
    customer_message["Subject"] = "MAWB# 98765432109"
    customer_message.set_content("customer body")
    customer_eml_path = tmp_path / "customer.eml"
    customer_eml_path.write_bytes(customer_message.as_bytes())
    preview = BookingPreview(
        session_id="sil-warehouse-service",
        supplier="SIL-FUCA",
        source_filename="customer.eml",
        pack_filename="",
        rows=[{"P/N": "TEST"}],
        columns=["P/N"],
        email_subject="MAWB# 98765432109",
    )

    output_path = booking_service.write_warehouse_mail(
        session_id=preview.session_id,
        preview=preview,
        customer_eml_path=customer_eml_path,
        warehouse_file=UploadFile(file=BytesIO(payload), filename=warehouse_filename),
    )

    generated = BytesParser(policy=policy.default).parsebytes(output_path.read_bytes())
    attachments = [part for part in generated.walk() if part.get_content_disposition() == "attachment"]
    assert [part.get_filename() for part in attachments] == [f"booking_warehouse_{warehouse_filename}"]
    assert attachments[0].get_content_type() == expected_content_type
    assert attachments[0].get_payload(decode=True) == payload


def test_write_sil_warehouse_mail_rejects_unlisted_attachment_type(monkeypatch, tmp_path: Path) -> None:
    monkeypatch.setattr(
        booking_service,
        "save_upload",
        partial(save_upload, upload_root=tmp_path / "uploads"),
    )

    with pytest.raises(UploadValidationError, match=r"allowed suffixes: .*\.pdf.*\.xls.*\.xlsx"):
        booking_service.write_warehouse_mail(
            session_id="sil-warehouse-invalid",
            preview=SimpleNamespace(supplier="SIL-FUCA"),
            customer_eml_path=tmp_path / "customer.eml",
            warehouse_file=UploadFile(file=BytesIO(b"not allowed"), filename="SIL26040490.docx"),
        )


def test_generate_sil_warehouse_eml_rejects_other_suppliers(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="只有 SIL-FUCA"):
        mail_builder.generate_sil_fuca_warehouse_eml(
            preview=SimpleNamespace(supplier="SIL-WEIKENG"),
            customer_eml_path=tmp_path / "customer.eml",
            warehouse_file_path=tmp_path / "SIL26040490.pdf",
            output_path=tmp_path / "warehouse.eml",
        )


def test_vc_dzyq_rule_maps_desktop_document_requirements(monkeypatch, tmp_path) -> None:
    source_path = tmp_path / "VC_DZYQ.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "清单 "
    ws.append(
        [
            "厂商编号",
            "厂商",
            "Customer_PO",
            "",
            "PO Item",
            "CustPartNumber",
            "CPNRev",
            "TI_MATERIAL",
            "BoxQty",
            "Net",
            "Net Wght Unit",
            "Total Wght",
            "Total Wght Unit",
            "PLS",
        ]
    )
    ws.append(["VC_DZYQ", "TI原厂", "C33C-26010025", "0001", "1010170933002T", "", "TLV70933PKR", 2000, 98, "G", 0.79, "KG", 2, "US"])
    ws.append([None, None, None, None, None, None, "合计", 2000, None, None, None, None, 2, None])
    wb.save(source_path)

    monkeypatch.setattr(vc_dzyq, "load_min_pack_lookup", lambda: ({"1010170933002T": 1000}, ""))

    preview = build_booking_preview(session_id="vc-test", supplier="VC_DZYQ", source_path=source_path)

    assert preview.can_generate
    assert len(preview.rows) == 1
    row = preview.rows[0]
    assert row["订单号"] == "C33C-26010025-0001"
    assert row["启益料号"] == "1010170933002T"
    assert row["商品名称"] == "IC"
    assert row["数量"] == 2000
    assert row["单位"] == "PCS"
    assert row["纸箱数"] == 2
    assert row["净重"] == 0.098
    assert row["毛重"] == 0.79
    assert row["体积"] == 0.01
    assert row["生产日期"] == date.today().strftime("%Y-%m-%d")
    assert row["产地 (made in)"] == "US"
    assert row["批次"] == "0"
    assert row["品牌"] == "无"
    assert row["LEDBinCode"] == "无"
    assert row["最小包装数"] == 1000
    assert row["每箱标准数"] == 2000


def test_booking_preview_maps_origin_to_country_abbreviation(monkeypatch, tmp_path) -> None:
    source_path = tmp_path / "VC_DZYQ_origin.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "清单 "
    ws.append(
        [
            "厂商编号",
            "厂商",
            "Customer_PO",
            "",
            "PO Item",
            "CustPartNumber",
            "CPNRev",
            "TI_MATERIAL",
            "BoxQty",
            "Net",
            "Net Wght Unit",
            "Total Wght",
            "Total Wght Unit",
            "PLS",
        ]
    )
    ws.append(["VC_DZYQ", "TI原厂", "C33C-26010025", "0001", "1010170933002T", "", "TLV70933PKR", 2000, 98, "G", 0.79, "KG", 2, "PHILIPPINES"])
    wb.save(source_path)

    monkeypatch.setattr(vc_dzyq, "load_min_pack_lookup", lambda: ({"1010170933002T": 1000}, ""))
    monkeypatch.setattr(booking_store, "load_country_abbr_lookup", lambda: ({"PHILIPPINES": "PH"}, ""))

    preview = build_booking_preview(session_id="origin-test", supplier="VC_DZYQ", source_path=source_path)

    assert preview.can_generate
    assert preview.rows[0]["产地 (made in)"] == "PH"


def test_write_booking_workbook_unmerges_template_rows_for_all_suppliers(tmp_path) -> None:
    columns = [
        "订单号",
        "启益料号",
        "商品名称",
        "数量",
        "单位",
        "纸箱数",
        "净重",
        "毛重",
        "体积",
        "产地 (made in)",
        "批次",
        "品牌",
        "LEDBinCode",
        "最小包装数",
        "每箱标准数",
    ]
    rows = [
        {
            "订单号": f"PO-{index:04d}",
            "启益料号": f"PART-{index:04d}",
            "商品名称": "IC",
            "数量": 1000 + index,
            "单位": "PCS",
            "纸箱数": index,
            "净重": index / 10,
            "毛重": index / 10 + 1,
            "体积": 0.01 if index == 1 else 0,
            "产地 (made in)": "US",
            "批次": "0",
            "品牌": "无",
            "LEDBinCode": "无",
            "最小包装数": 1000 + index,
            "每箱标准数": 1000 + index,
        }
        for index in range(1, 17)
    ]
    preview = BookingPreview(
        session_id="global-merge-test",
        supplier="SIL-FUCA",
        source_filename="synthetic.xlsx",
        pack_filename="",
        rows=rows,
        columns=columns,
    )
    output_path = write_booking_workbook(preview, tmp_path)
    output_wb = load_workbook(output_path, data_only=True)
    output_ws = output_wb[output_wb.sheetnames[0]]

    assert output_ws["C20"].value == "PO-0012"
    assert output_ws["D20"].value == "PART-0012"
    assert output_ws["E20"].value == "IC"
    assert output_ws["C21"].value == "PO-0013"
    assert output_ws["D21"].value == "PART-0013"
    assert output_ws["C24"].value == "PO-0016"
    assert output_ws["D24"].value == "PART-0016"
    assert output_ws["B20"].style_id == output_ws["B19"].style_id
    assert output_ws["C20"].style_id == output_ws["C19"].style_id
    assert output_ws["E20"].style_id == output_ws["E19"].style_id
    assert output_ws["B21"].style_id == output_ws["B19"].style_id
    assert output_ws["C21"].style_id == output_ws["C19"].style_id
    assert output_ws["E24"].style_id == output_ws["E19"].style_id
    assert output_ws["A25"].value == "Total"
    assert output_ws["A26"].value == "* required field"
    assert "A25:E25" in {str(item) for item in output_ws.merged_cells.ranges}
    assert "A26:C26" in {str(item) for item in output_ws.merged_cells.ranges}
