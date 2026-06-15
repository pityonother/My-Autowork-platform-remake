from __future__ import annotations

from openpyxl import Workbook

import export_clearance_store as store
from app.modules.export_clearance.rules import (
    export_record_business_key,
    format_pallet_carton_text,
    is_tan_number,
    normalize_tan_number,
    urgency_sort_key,
)


def test_export_clearance_rule_helpers_normalize_display_values() -> None:
    assert normalize_tan_number("tan#123") == "TAN#123"
    assert normalize_tan_number("TAN84273") == "TAN#84273"
    assert is_tan_number("TAN84273")
    assert is_tan_number("tan#84274")
    assert not is_tan_number("TAN#")
    assert format_pallet_carton_text("2.0", "3.0") == "2板/3箱"


def test_export_clearance_parser_accepts_tan_without_hash(tmp_path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["装货明细"])
    sheet.append(["Item", "SN NO.", "ship mode", "目的国", "总数量PCS", "总价USD", "毛重KG", "箱数", "卡板数"])
    sheet.append([1, "SN-1", "by sea", "FR", 10, 100, 20, 10, 1])
    sheet.append(["TAN100", "first"])
    sheet.append([2, "SN-2", "by sea", "DE", 5, 50, 7, 5, 1])
    sheet.append([None, "SN-2", "by sea", "DE", 6, 60, 8, 6, 0])
    sheet.append(["tan#101", "second"])
    source_path = tmp_path / "clearance.xlsx"
    workbook.save(source_path)

    records = store.parse_export_clearance_file(source_path)

    assert [record["tan_number"] for record in records] == ["TAN#100", "TAN#101"]
    assert records[1]["total_pcs"] == 11
    assert records[1]["carton_count"] == 11


def test_export_clearance_parser_uses_item_column_when_sheet_has_left_padding(tmp_path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["装货明细"])
    sheet.append(["", "Item", "SN NO.", "ship mode", "目的国", "总数量PCS", "总价USD", "毛重KG", "箱数", "卡板数"])
    sheet.append(["", 1, "SN-1", "by air", "DK", 4, 40, 8, 4, 1])
    sheet.append(["", "TAN#84305", "first"])
    source_path = tmp_path / "clearance_left_padding.xlsx"
    workbook.save(source_path)

    records = store.parse_export_clearance_file(source_path)

    assert len(records) == 1
    assert records[0]["tan_number"] == "TAN#84305"
    assert records[0]["tan_description"] == "first"
    assert records[0]["sn_number"] == "SN-1"
    assert records[0]["total_pcs"] == 4


def test_export_clearance_parser_accepts_tan_note_in_sn_column(tmp_path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["装货明细"])
    sheet.append(["Item", "SN NO.", "ship mode", "目的国", "总数量PCS", "总价USD", "毛重KG", "箱数", "卡板数"])
    sheet.append([1, "SN-1", "by air", "DK", 4, 40, 8, 4, 1])
    sheet.append(["", "TAN#84305", "first"])
    source_path = tmp_path / "clearance_tan_in_sn_column.xlsx"
    workbook.save(source_path)

    records = store.parse_export_clearance_file(source_path)

    assert len(records) == 1
    assert records[0]["tan_number"] == "TAN#84305"
    assert records[0]["tan_description"] == "first"
    assert records[0]["sn_number"] == "SN-1"
    assert records[0]["total_pcs"] == 4


def test_export_record_business_key_uses_core_identity_fields() -> None:
    record = {
        "tan_number": "TAN#1",
        "total_pcs": "10",
        "carton_count": "2",
        "pallet_count": "1",
        "total_value_usd": "33.5",
        "gross_weight_kg": "44.2",
    }

    assert export_record_business_key(" M1 ", "2026-05-18", record) == (
        "M1",
        "2026-05-18",
        "TAN#1",
        10,
        2,
        1,
        33.5,
        44.2,
    )


def test_export_clearance_urgency_sort_orders_red_first() -> None:
    rows = [
        {"urgency_color": "white", "shipment_date": "2026-05-18", "tan_number": "TAN#2"},
        {"urgency_color": "red", "shipment_date": "2026-05-17", "tan_number": "TAN#1"},
        {"urgency_color": "orange", "shipment_date": "2026-05-16", "tan_number": "TAN#3"},
    ]

    assert [item["urgency_color"] for item in sorted(rows, key=urgency_sort_key)] == ["red", "orange", "white"]
